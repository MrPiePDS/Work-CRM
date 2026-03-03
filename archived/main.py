# -*- coding: utf-8 -*-
import sys
import json
import csv
import hashlib
import traceback
import html as _html
import subprocess
import tempfile
import zipfile
import os
from pathlib import Path
from datetime import datetime, date
from pathlib import Path
from typing import Optional, List, Tuple
import urllib.request

from PySide6.QtCore import Qt, QMarginsF, QRegularExpression
from PySide6.QtGui import QTextDocument, QRegularExpressionValidator, QPageLayout, QPageSize
from PySide6.QtPrintSupport import QPrinter
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit,
    QPushButton, QMessageBox, QFrame, QComboBox, QCheckBox, QTabWidget,
    QTableWidget, QTableWidgetItem, QDialog, QTextEdit, QFileDialog,
    QListWidget, QListWidgetItem, QScrollArea, QSizePolicy, QGridLayout,
    QSplitter, QAbstractItemView, QHeaderView, QProgressBar
)

from openpyxl import Workbook, load_workbook


# ----------------------------- Data Paths -----------------------------
APP_NAME = "ClientManagerV1"

BOOT_DIR = Path.home() / "AppData" / "Roaming" / APP_NAME
BOOT_DIR.mkdir(parents=True, exist_ok=True)
BOOT_USERS_PATH = BOOT_DIR / "users.json"

def read_boot_data_root() -> str:
    try:
        if BOOT_USERS_PATH.exists():
            users = json.loads(BOOT_USERS_PATH.read_text(encoding="utf-8"))
            return (users.get("_settings", {}) or {}).get("data_root", "") or ""
    except Exception:
        pass
    return ""


def data_dir() -> Path:
    base = Path(os.getenv("APPDATA") or Path.home())
    app = base / "ClientManagerV1"
    app.mkdir(parents=True, exist_ok=True)
    return app

DATA_DIR = data_dir()
EXCEL_PATH = DATA_DIR / "clients.xlsx"
USERS_PATH = DATA_DIR / "users.json"
AUDIT_PATH = DATA_DIR / "audit_log.csv"

DEFAULT_CLIENTS_ROOT = DATA_DIR / "clients"
DEFAULT_CLIENTS_ROOT.mkdir(parents=True, exist_ok=True)


# ----------------------------- Improved PDF Printing -----------------------------
def ensure_sumatrapdf() -> Optional[Path]:
    """Ελέγχει και κατεβάζει το SumatraPDF αν δεν υπάρχει."""
    sumatra_exe = DATA_DIR / "SumatraPDF.exe"
    if sumatra_exe.exists():
        return sumatra_exe

    try:
        url = "https://www.sumatrapdfreader.org/dl/rel/3.5.2/SumatraPDF-3.5.2-64.zip"

        QMessageBox.information(
            None,
            "Λήψη SumatraPDF",
            "Το πρόγραμμα θα κατεβάσει το SumatraPDF (portable) για σωστή εκτύπωση PDF.\n"
            "Αυτό γίνεται μόνο μια φορά.\n\n"
            "Παρακαλώ περιμένετε..."
        )

        temp_zip = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
        temp_zip_path = temp_zip.name
        temp_zip.close()

        try:
            urllib.request.urlretrieve(url, temp_zip_path)

            with zipfile.ZipFile(temp_zip_path, "r") as zip_ref:
                extracted = None
                for info in zip_ref.infolist():
                    if info.filename.endswith("SumatraPDF.exe") or "SumatraPDF.exe" in info.filename:
                        zip_ref.extract(info, DATA_DIR)
                        extracted = DATA_DIR / info.filename
                        break

            Path(temp_zip_path).unlink(missing_ok=True)

            if extracted and extracted.exists():
                target = DATA_DIR / "SumatraPDF.exe"
                try:
                    if extracted.resolve() != target.resolve():
                        target.parent.mkdir(parents=True, exist_ok=True)
                        extracted.replace(target)
                    return target
                except Exception:
                    return extracted

        except Exception as e:
            print(f"Σφάλμα λήψης/αποσυμπίεσης: {e}")
            return None

    except Exception as e:
        print(f"Αποτυχία λήψης SumatraPDF: {e}")
        return None

    return None


def print_pdf_with_sumatra(pdf_path: Path) -> bool:
    """Εκτύπωση PDF με SumatraPDF (πιο αξιόπιστη)."""
    sumatra = ensure_sumatrapdf()
    if sumatra and sumatra.exists():
        try:
            cmd = [str(sumatra), "-print-to-default", "-silent", str(pdf_path)]
            subprocess.run(cmd, timeout=30, check=True)
            return True
        except Exception as e:
            print(f"Σφάλμα εκτύπωσης με SumatraPDF: {e}")
            return print_file_windows(pdf_path)
    return print_file_windows(pdf_path)


def print_file_windows(path: Path) -> bool:
    """Fallback εκτύπωση/άνοιγμα με Windows."""
    try:
        import os
        if path.suffix.lower() == ".pdf":
            sumatra = ensure_sumatrapdf()
            if sumatra and sumatra.exists():
                return print_pdf_with_sumatra(path)

        os.startfile(str(path), "print")
        return True
    except Exception as e:
        print(f"Σφάλμα εκτύπωσης: {e}")
        return False


def print_with_dialog(path: Path, parent=None) -> bool:
    """Εκτύπωση αρχείου. Για PDF → Sumatra, αλλιώς άνοιγμα."""
    try:
        import os
        if path.suffix.lower() == ".pdf":
            return print_pdf_with_sumatra(path)
        os.startfile(str(path))
        return True
    except Exception as e:
        print(f"Σφάλμα εκτύπωσης με διάλογο: {e}")
        return print_file_windows(path)


# ----------------------------- Styles (POLISHED) -----------------------------
UI = {
    "radius_card": 14,
    "radius_sub": 12,
    "radius_input": 10,
    "radius_btn": 12,
    "pad_card": 10,
    "pad_input_v": 7,
    "pad_input_h": 10,
    "pad_btn_v": 9,
    "pad_btn_h": 12,
}

DARK_STYLE = f"""
* {{
    font-family: "Segoe UI";
    font-size: 13px;
}}
QWidget {{
    background: #0F1012;
    color: #E7E3E8;
}}
/* Cards */
QFrame#card {{
    background: #17181B;
    border: 1px solid #262833;
    border-radius: {UI["radius_card"]}px;
}}
QFrame#subcard {{
    background: #141519;
    border: 1px solid #262833;
    border-radius: {UI["radius_sub"]}px;
}}
/* Text */
QLabel#title {{
    font-size: 20px;
    font-weight: 650;
}}
QLabel#subtitle {{
    font-size: 14px;
    font-weight: 650;
}}
QLabel#muted {{
    color: #AAA7B0;
}}
QLabel#chip {{
    background: #1C1D22;
    border: 1px solid #2A2C36;
    border-radius: 10px;
    padding: 5px 10px;
}}
QLabel#chipGood {{
    background: #15301E;
    border: 1px solid #2E6E3F;
    border-radius: 10px;
    padding: 5px 10px;
}}
QLabel#chipBad {{
    background: #2E1414;
    border: 1px solid #7A2B2B;
    border-radius: 10px;
    padding: 5px 10px;
}}
/* Inputs */
QLineEdit, QTextEdit, QComboBox {{
    background: #1B1C21;
    border: 1px solid #2A2C36;
    border-radius: {UI["radius_input"]}px;
    padding: {UI["pad_input_v"]}px {UI["pad_input_h"]}px;
}}
QLineEdit:focus, QTextEdit:focus, QComboBox:focus {{
    border: 2px solid #7F67BE;
}}
QLineEdit:disabled, QTextEdit:disabled, QComboBox:disabled {{
    color: #8A8790;
    background: #141519;
    border: 1px solid #232532;
}}
QLineEdit::placeholder {{
    color: #7E7A86;
}}
/* Combo popup */
QComboBox::drop-down {{
    subcontrol-origin: padding;
    subcontrol-position: top right;
    width: 28px;
    border-left: 1px solid #2A2C36;
}}
QComboBox QAbstractItemView {{
    background: #17181B;
    border: 1px solid #2A2C36;
    outline: 0px;
    selection-background-color: #2B2550;
}}
/* Buttons */
QPushButton {{
    background: #7F67BE;
    border: 1px solid #5F4CA6;
    color: #FFFFFF;
    border-radius: {UI["radius_btn"]}px;
    padding: {UI["pad_btn_v"]}px {UI["pad_btn_h"]}px;
    font-weight: 650;
}}
QPushButton:hover {{
    background: #8F79CC;
}}
QPushButton:pressed {{
    background: #6F58B0;
}}
QPushButton:disabled {{
    background: #2A2C36;
    border: 1px solid #2A2C36;
    color: #8A8790;
}}
QPushButton#secondary {{
    background: #1B1C21;
    border: 1px solid #2A2C36;
    color: #E7E3E8;
}}
QPushButton#secondary:hover {{
    background: #20222A;
}}
/* Drop zone */
QLabel#drop {{
    border: 2px dashed #7F67BE;
    border-radius: {UI["radius_card"]}px;
    padding: 14px;
    color: #CBBEF5;
    background: rgba(127,103,190,0.08);
}}
/* Tabs */
QTabWidget::pane {{
    border: 0px;
}}
QTabBar::tab {{
    background: #17181B;
    border: 1px solid #262833;
    border-bottom: 0px;
    padding: 9px 14px;
    border-top-left-radius: 10px;
    border-top-right-radius: 10px;
    margin-right: 6px;
    min-width: 110px;
}}
QTabBar::tab:selected {{
    background: #1D1F26;
}}
/* Tables */
QTableWidget {{
    background: #17181B;
    border: 1px solid #262833;
    border-radius: 12px;
    gridline-color: #262833;
    selection-background-color: #2B2550;
    selection-color: #FFFFFF;
}}
QTableWidget::item {{
    padding: 6px 8px;
    border: 0px;
}}
QHeaderView::section {{
    background: #1D1F26;
    color: #E7E3E8;
    padding: 8px 8px;
    border: 0px;
    border-bottom: 1px solid #262833;
    font-weight: 650;
}}
QTableCornerButton::section {{
    background: #1D1F26;
    border: 0px;
}}
/* Lists */
QListWidget {{
    background: #17181B;
    border: 1px solid #262833;
    border-radius: 12px;
}}
QListWidget::item {{
    padding: 6px 8px;
    border-bottom: 1px solid #232532;
}}
QListWidget::item:selected {{
    background: #2B2550;
}}
/* Checkboxes */
QCheckBox {{
    spacing: 8px;
}}
QCheckBox#themeSwitch::indicator {{
    width: 46px;
    height: 24px;
    border-radius: 12px;
    background: #2A2C36;
    border: 1px solid #3A3C46;
}}
QCheckBox#themeSwitch::indicator:checked {{
    background: #7F67BE;
    border: 1px solid #8F79CC;
}}
/* Scrollbars */
QScrollBar:vertical {{
    border: none;
    background: transparent;
    width: 10px;
    margin: 2px;
}}
QScrollBar::handle:vertical {{
    background: #7F67BE;
    border-radius: 5px;
    min-height: 24px;
}}
QScrollBar::handle:vertical:hover {{
    background: #8F79CC;
}}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
    border: none;
    background: none;
    height: 0px;
}}
QScrollBar:horizontal {{
    border: none;
    background: transparent;
    height: 10px;
    margin: 2px;
}}
QScrollBar::handle:horizontal {{
    background: #7F67BE;
    border-radius: 5px;
    min-width: 24px;
}}
QScrollBar::handle:horizontal:hover {{
    background: #8F79CC;
}}
/* Tooltips */
QToolTip {{
    background: #1D1F26;
    color: #E7E3E8;
    border: 1px solid #262833;
    padding: 6px 8px;
    border-radius: 10px;
}}
/* Progress Bar */
QProgressBar {{
    border: 1px solid #2A2C36;
    border-radius: 8px;
    text-align: center;
    color: #E7E3E8;
    background: #1B1C21;
}}
QProgressBar::chunk {{
    border-radius: 8px;
}}
"""

LIGHT_STYLE = f"""
* {{
    font-family: "Segoe UI";
    font-size: 13px;
}}
QWidget {{
    background: #F7F7FB;
    color: #1A1A1A;
}}
/* Cards */
QFrame#card {{
    background: #FFFFFF;
    border: 1px solid #E2E2EE;
    border-radius: {UI["radius_card"]}px;
}}
QFrame#subcard {{
    background: #FFFFFF;
    border: 1px solid #E2E2EE;
    border-radius: {UI["radius_sub"]}px;
}}
/* Text */
QLabel#title {{
    font-size: 20px;
    font-weight: 650;
}}
QLabel#subtitle {{
    font-size: 14px;
    font-weight: 650;
}}
QLabel#muted {{
    color: #5B5D66;
}}
QLabel#chip {{
    background: #F2F2F7;
    border: 1px solid #E2E2EE;
    border-radius: 10px;
    padding: 5px 10px;
}}
QLabel#chipGood {{
    background: #E7F6EA;
    border: 1px solid #BFE7C8;
    border-radius: 10px;
    padding: 5px 10px;
}}
QLabel#chipBad {{
    background: #FDEAEA;
    border: 1px solid #F2BCBC;
    border-radius: 10px;
    padding: 5px 10px;
}}
/* Inputs */
QLineEdit, QTextEdit, QComboBox {{
    background: #FFFFFF;
    border: 1px solid #D6D6E2;
    border-radius: {UI["radius_input"]}px;
    padding: {UI["pad_input_v"]}px {UI["pad_input_h"]}px;
}}
QLineEdit:focus, QTextEdit:focus, QComboBox:focus {{
    border: 2px solid #6750A4;
}}
QLineEdit:disabled, QTextEdit:disabled, QComboBox:disabled {{
    color: #8A8C96;
    background: #F2F2F7;
    border: 1px solid #E2E2EE;
}}
QLineEdit::placeholder {{
    color: #8A8C96;
}}
/* Combo popup */
QComboBox::drop-down {{
    subcontrol-origin: padding;
    subcontrol-position: top right;
    width: 28px;
    border-left: 1px solid #D6D6E2;
}}
QComboBox QAbstractItemView {{
    background: #FFFFFF;
    border: 1px solid #D6D6E2;
    outline: 0px;
    selection-background-color: #E8E2FF;
}}
/* Buttons */
QPushButton {{
    background: #6750A4;
    border: 1px solid #56408F;
    color: #FFFFFF;
    border-radius: {UI["radius_btn"]}px;
    padding: {UI["pad_btn_v"]}px {UI["pad_btn_h"]}px;
    font-weight: 650;
}}
QPushButton:hover {{
    background: #7F67BE;
}}
QPushButton:pressed {{
    background: #5D4698;
}}
QPushButton:disabled {{
    background: #E2E2EE;
    border: 1px solid #E2E2EE;
    color: #8A8C96;
}}
QPushButton#secondary {{
    background: #F2F2F7;
    border: 1px solid #E2E2EE;
    color: #1A1A1A;
}}
QPushButton#secondary:hover {{
    background: #E9E9F2;
}}
/* Drop zone */
QLabel#drop {{
    border: 2px dashed #6750A4;
    border-radius: {UI["radius_card"]}px;
    padding: 14px;
    color: #3B2A6B;
    background: rgba(103,80,164,0.06);
}}
/* Tabs */
QTabWidget::pane {{
    border: 0px;
}}
QTabBar::tab {{
    background: #FFFFFF;
    border: 1px solid #E2E2EE;
    border-bottom: 0px;
    padding: 9px 14px;
    border-top-left-radius: 10px;
    border-top-right-radius: 10px;
    margin-right: 6px;
    min-width: 110px;
}}
QTabBar::tab:selected {{
    background: #F2F2F7;
}}
/* Tables */
QTableWidget {{
    background: #FFFFFF;
    border: 1px solid #E2E2EE;
    border-radius: 12px;
    gridline-color: #E2E2EE;
    selection-background-color: #E8E2FF;
    selection-color: #1A1A1A;
}}
QTableWidget::item {{
    padding: 6px 8px;
    border: 0px;
}}
QHeaderView::section {{
    background: #F2F2F7;
    color: #1A1A1A;
    padding: 8px 8px;
    border: 0px;
    border-bottom: 1px solid #E2E2EE;
    font-weight: 650;
}}
QTableCornerButton::section {{
    background: #F2F2F7;
    border: 0px;
}}
/* Lists */
QListWidget {{
    background: #FFFFFF;
    border: 1px solid #E2E2EE;
    border-radius: 12px;
}}
QListWidget::item {{
    padding: 6px 8px;
    border-bottom: 1px solid #EDEDF6;
}}
QListWidget::item:selected {{
    background: #E8E2FF;
}}
/* Checkboxes */
QCheckBox {{
    spacing: 8px;
}}
QCheckBox#themeSwitch::indicator {{
    width: 46px;
    height: 24px;
    border-radius: 12px;
    background: #D6D6E2;
    border: 1px solid #C7C7D4;
}}
QCheckBox#themeSwitch::indicator:checked {{
    background: #6750A4;
    border: 1px solid #7F67BE;
}}
/* Scrollbars */
QScrollBar:vertical {{
    border: none;
    background: transparent;
    width: 10px;
    margin: 2px;
}}
QScrollBar::handle:vertical {{
    background: #6750A4;
    border-radius: 5px;
    min-height: 24px;
}}
QScrollBar::handle:vertical:hover {{
    background: #7F67BE;
}}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
    border: none;
    background: none;
    height: 0px;
}}
QScrollBar:horizontal {{
    border: none;
    background: transparent;
    height: 10px;
    margin: 2px;
}}
QScrollBar::handle:horizontal {{
    background: #6750A4;
    border-radius: 5px;
    min-width: 24px;
}}
QScrollBar::handle:horizontal:hover {{
    background: #7F67BE;
}}
/* Tooltips */
QToolTip {{
    background: #FFFFFF;
    color: #1A1A1A;
    border: 1px solid #E2E2EE;
    padding: 6px 8px;
    border-radius: 10px;
}}
/* Progress Bar */
QProgressBar {{
    border: 1px solid #D6D6E2;
    border-radius: 8px;
    text-align: center;
    color: #1A1A1A;
    background: #FFFFFF;
}}
QProgressBar::chunk {{
    border-radius: 8px;
}}
"""


def polish_table(t: QTableWidget) -> None:
    """UI-only tweaks for cleaner tables."""
    t.setAlternatingRowColors(True)
    t.setShowGrid(False)
    t.verticalHeader().setVisible(False)
    t.horizontalHeader().setStretchLastSection(False)
    t.setSortingEnabled(False)


def polish_list(w: QListWidget) -> None:
    """UI-only tweaks for cleaner lists."""
    w.setAlternatingRowColors(True)


# ----------------------------- Excel Schema -----------------------------
SHEET_NAME = "Clients"
HEADERS = [
    "ID", "ServiceType", "Date", "Name", "Phone", "AFM", "AMKA", "AMA",
    "HasTaxisnet", "TaxisnetUser", "TaxisnetPass", "Kleidarithmos", "AMKA_AMA",
    "Aporipsi", "ActionsToday", "PaymentMethod", "RequestNotes", "Total", "Paid",
    "Balance", "FolderPath", "FilesConfirmedBy", "FilesConfirmedAt", "CreatedBy",
    "CreatedAt", "LastEditedBy", "LastEditedAt",
    # Νέα πεδία για ενίσχυση καρτέλας
    "Goal", "DeclarationStatus", "CustomerStatus", "Amount", "AMKA_Valid"
]


SERVICES = [
    ("ΑΜΚΑ / ΑΜΑ", 160),
    ("Μεταβολή", 20),
    ("Κλειδάριθμος", 20),
    ("ΑΦΜ", 50),
    ("Εργασία", 75),
]

PAYMENT_METHODS = ["(κανένα)", "Μετρητά", "Κάρτα", "Τραπεζική μεταφορά"]


# ----------------------------- Core Utils -----------------------------
def safe_float_str(s: str) -> float:
    return safe_float(s)


def read_audit_for_customer(customer_id: int) -> List[List[str]]:
    if not AUDIT_PATH.exists():
        return []
    rows: List[List[str]] = []
    with AUDIT_PATH.open("r", encoding="utf-8") as f:
        r = csv.reader(f)
        _ = next(r, None)
        for line in r:
            try:
                cid = str(line[3]).strip()
                if cid == str(customer_id):
                    rows.append(line)
            except Exception:
                pass
    return rows


# ----------------------------- Storage: Users + Settings -----------------------------
def ensure_users_file() -> None:
    if USERS_PATH.exists():
        return
    users = {
        "_settings": {
            "scanner_folder": "",
            "theme": "light",  # LIGHT default
            "clients_root": str(DEFAULT_CLIENTS_ROOT),
        },
        "admin": {"password_hash": sha256("1234"), "created_at": now_iso()}
    }
    USERS_PATH.write_text(json.dumps(users, indent=2, ensure_ascii=False), encoding="utf-8")


def load_users() -> dict:
    ensure_users_file()
    return json.loads(USERS_PATH.read_text(encoding="utf-8"))


def save_users(users: dict) -> None:
    USERS_PATH.write_text(json.dumps(users, indent=2, ensure_ascii=False), encoding="utf-8")


def verify_user(username: str, password: str) -> bool:
    users = load_users()
    u = users.get(username)
    if not u:
        return False
    return u.get("password_hash") == sha256(password)


def set_user_password(username: str, new_password: str) -> None:
    users = load_users()
    if username not in users:
        users[username] = {"created_at": now_iso()}
    users[username]["password_hash"] = sha256(new_password)
    save_users(users)


def get_scanner_folder() -> str:
    users = load_users()
    return (users.get("_settings", {}) or {}).get("scanner_folder", "") or ""


def set_scanner_folder(path: str) -> None:
    users = load_users()
    users.setdefault("_settings", {})
    users["_settings"]["scanner_folder"] = path
    save_users(users)


def get_theme() -> str:
    users = load_users()
    t = (users.get("_settings", {}) or {}).get("theme", "light")
    return t if t in ("dark", "light") else "light"


def set_theme(theme: str) -> None:
    users = load_users()
    users.setdefault("_settings", {})
    users["_settings"]["theme"] = "light" if theme == "light" else "dark"
    save_users(users)


def get_clients_root() -> Path:
    users = load_users()
    raw = (users.get("_settings", {}) or {}).get("clients_root", "") or ""
    p = Path(raw) if raw else DEFAULT_CLIENTS_ROOT
    try:
        p.mkdir(parents=True, exist_ok=True)
    except Exception:
        p = DEFAULT_CLIENTS_ROOT
        p.mkdir(parents=True, exist_ok=True)
    return p


def set_clients_root(path: str) -> None:
    users = load_users()
    users.setdefault("_settings", {})
    users["_settings"]["clients_root"] = path
    save_users(users)

def get_data_root() -> str:
    users = load_users()
    return (users.get("_settings", {}) or {}).get("data_root", "") or ""

def set_data_root(path: str) -> None:
    users = load_users()
    users.setdefault("_settings", {})
    users["_settings"]["data_root"] = path
    save_users(users)



# ----------------------------- Storage: Excel (with migration) -----------------------------
def ensure_excel() -> None:
    try:
        if not EXCEL_PATH.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = SHEET_NAME
            ws.append(HEADERS)
            wb.save(EXCEL_PATH)
            return

        wb = load_workbook(EXCEL_PATH)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.active
            ws.title = SHEET_NAME
            if ws.max_row < 1:
                ws.append(HEADERS)

        existing = [str(c.value or "").strip() for c in ws[1]] if ws.max_row >= 1 else []
        changed = False
        for h in HEADERS:
            if h not in existing:
                existing.append(h)
                ws.cell(row=1, column=len(existing)).value = h
                changed = True
        if changed:
            wb.save(EXCEL_PATH)

    except Exception:
        try:
            backup = EXCEL_PATH.with_name(f"clients_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            if EXCEL_PATH.exists():
                EXCEL_PATH.replace(backup)
        except Exception:
            pass

        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        wb.save(EXCEL_PATH)


def open_ws():
    ensure_excel()
    try:
        wb = load_workbook(EXCEL_PATH)
    except Exception as e:
        raise RuntimeError(
            "Δεν μπορώ να ανοίξω το clients.xlsx.\n"
            "Κλείσε το Excel αν είναι ανοιχτό και ξαναδοκίμασε.\n\n"
            f"Λεπτομέρειες: {e}"
        )

    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    if ws.title != SHEET_NAME:
        ws.title = SHEET_NAME
    return wb, ws


def headers_map(ws) -> dict:
    cols = {}
    for i, cell in enumerate(ws[1], start=1):
        cols[str(cell.value or "").strip()] = i
    return cols


def append_record_by_headers(ws, data: dict) -> int:
    """
    Προσθέτει νέα γραμμή γράφοντας τιμές με βάση ΤΑ ΟΝΟΜΑΤΑ στηλών,
    όχι τη σειρά στηλών στο Excel.
    """
    cols = headers_map(ws)

    # ensure all headers exist
    for h in HEADERS:
        if h not in cols:
            ws.cell(row=1, column=ws.max_column + 1).value = h

    cols = headers_map(ws)
    new_r = ws.max_row + 1

    for k, v in data.items():
        if k in cols:
            ws.cell(row=new_r, column=cols[k]).value = v

    return new_r


def get_next_id() -> int:
    wb, ws = open_ws()
    cols = headers_map(ws)
    id_col = cols.get("ID", 1)

    used = set()
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, id_col).value
        try:
            i = int(str(v).strip())
            if i > 0:
                used.add(i)
        except Exception:
            pass

    nxt = 1
    while nxt in used:
        nxt += 1
    return nxt


def find_rows(query_id: str = "", query_name: str = "", query_phone: str = "") -> List[int]:
    wb, ws = open_ws()
    cols = headers_map(ws)

    query_id = query_id.strip()
    query_name = query_name.strip().lower()
    query_phone = query_phone.strip()

    results = []
    for r in range(2, ws.max_row + 1):
        cid = str(ws.cell(r, cols.get("ID", 1)).value or "").strip()
        name = str(ws.cell(r, cols.get("Name", 1)).value or "").strip().lower()
        phone = str(ws.cell(r, cols.get("Phone", 1)).value or "").strip()

        ok = True
        if query_id:
            ok = ok and (cid == query_id)
        if query_name:
            ok = ok and (query_name in name)
        if query_phone:
            ok = ok and (query_phone in phone)

        if ok and (query_id or query_name or query_phone):
            results.append(r)
    return results


def row_to_record(ws, r: int) -> dict:
    cols = headers_map(ws)
    rec = {}
    for h, c in cols.items():
        rec[h] = ws.cell(r, c).value
    return rec


def update_row(ws, r: int, updates: dict) -> None:
    cols = headers_map(ws)
    for k, v in updates.items():
        if k in cols:
            ws.cell(r, cols[k]).value = v


# ----------------------------- Utilities -----------------------------
def now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def safe_float(s: str) -> float:
    s = (s or "").strip().replace(",", ".")
    if not s:
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def sha256(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


def audit_log(user: str, action: str, customer_id: Optional[int] = None, details: str = "") -> None:
    try:
        file_exists = AUDIT_PATH.exists()
        with AUDIT_PATH.open("a", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            if not file_exists:
                w.writerow(["timestamp", "user", "action", "customer_id", "details"])
            w.writerow([now_iso(), user, action, customer_id or "", details])
    except Exception:
        pass


def human_duration(from_iso: str) -> str:
    try:
        dt = datetime.fromisoformat(str(from_iso))
        diff = datetime.now() - dt
        mins = int(diff.total_seconds() // 60)
        hrs = mins // 60
        days = hrs // 24
        if days > 0:
            return f"{days}μ {hrs % 24}ω"
        if hrs > 0:
            return f"{hrs}ω {mins % 60}λ"
        return f"{mins}λ"
    except Exception:
        return "-"


# ----------------------------- Files -----------------------------
def sanitize_folder_part(name: str) -> str:
    safe = "".join(ch for ch in (name or "") if ch.isalnum() or ch in (" ", "_", "-", ".", "(", ")")).strip()
    safe = safe[:60] if safe else "Πελάτης"
    return safe


def customer_folder_name(customer_id: int, name: str) -> str:
    return f"{customer_id} - {sanitize_folder_part(name)}"


def resolve_customer_folder_from_record(rec: dict) -> Optional[Path]:
    """
    1) Αν το FolderPath υπάρχει και υπάρχει στο δίσκο -> use it
    2) Αλλιώς φτιάξε fallback: clients_root / "{id} - {name}"
    """
    fp_raw = str(rec.get("FolderPath") or "").strip()
    if fp_raw:
        p = Path(fp_raw)
        if p.exists():
            return p

    # fallback
    try:
        cid = int(rec.get("ID") or 0)
    except Exception:
        cid = 0
    name = str(rec.get("Name") or "").strip()

    if cid > 0:
        p2 = get_clients_root() / customer_folder_name(cid, name)
        if p2.exists():
            return p2.resolve()

    return None


def ensure_customer_folder(customer_id: int, name: str) -> Path:
    root = get_clients_root()
    folder = root / customer_folder_name(customer_id, name)
    folder.mkdir(parents=True, exist_ok=True)
    return folder.resolve()


def open_in_explorer(path: Path) -> None:
    try:
        import os
        p = Path(path)
        if p.exists():
            os.startfile(str(p))
    except Exception:
        pass


# ----------------------------- Validators / Limits -----------------------------
def set_digits_only(line: QLineEdit, max_len: int, allow_empty=True):
    line.setMaxLength(max_len)
    rx = QRegularExpression(r"^\d{0," + str(max_len) + r"}$")
    line.setValidator(QRegularExpressionValidator(rx))
    if not allow_empty:
        pass


def set_phone_validator(line: QLineEdit, max_len: int = 15):
    line.setMaxLength(max_len)
    rx = QRegularExpression(r"^[0-9+\-\s]{0," + str(max_len) + r"}$")
    line.setValidator(QRegularExpressionValidator(rx))


# ----------------------------- Custom Scroll Area -----------------------------
class ResponsiveScrollArea(QScrollArea):
    """ScrollArea που δεν έχει ποτέ horizontal scroll και είναι responsive."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWidgetResizable(True)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.setFrameShape(QFrame.NoFrame)

    def setWidget(self, widget):
        super().setWidget(widget)
        widget.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)


# ----------------------------- Scrollable Dialog -----------------------------
class ScrollableDialog(QDialog):
    """Dialog με scroll area για περιεχόμενο που δεν χωράει στην οθόνη."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        self.scroll_area = ResponsiveScrollArea()
        self.scroll_area.setWidgetResizable(True)

        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(20, 20, 20, 20)
        self.content_layout.setSpacing(15)

        self.scroll_area.setWidget(self.content_widget)
        main_layout.addWidget(self.scroll_area)

    def addWidget(self, widget):
        self.content_layout.addWidget(widget)

    def addLayout(self, layout):
        self.content_layout.addLayout(layout)

    def addStretch(self, stretch=0):
        self.content_layout.addStretch(stretch)


# ----------------------------- UI Components -----------------------------
class DropZone(QLabel):
    def __init__(self, on_files_dropped):
        super().__init__("📂 Σύρε & άφησε αρχεία εδώ")
        self.setObjectName("drop")
        self.setAlignment(Qt.AlignCenter)
        self.setAcceptDrops(True)
        self.on_files_dropped = on_files_dropped

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()

    def dropEvent(self, event):
        urls = event.mimeData().urls()
        paths = [u.toLocalFile() for u in urls if u.isLocalFile()]
        if paths:
            self.on_files_dropped(paths)


class CustomerCardWidget(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("card")

        self._excel_row_index: Optional[int] = None
        self._customer_id: Optional[int] = None
        self._folder_path: Optional[Path] = None

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(8)

        self.lbl_title = QLabel("Επέλεξε πελάτη")
        self.lbl_title.setObjectName("title")

        self.lbl_sub = QLabel("Κάνε κλικ σε γραμμή στον πίνακα αριστερά για προεπισκόπηση.")
        self.lbl_sub.setObjectName("muted")

        root.addWidget(self.lbl_title)
        root.addWidget(self.lbl_sub)

        chips_row = QHBoxLayout()
        chips_row.setSpacing(6)
        self.ch_id = QLabel("ID: -"); self.ch_id.setObjectName("chip")
        self.ch_balance = QLabel("Υπόλοιπο: -"); self.ch_balance.setObjectName("chip")
        self.ch_files = QLabel("Αρχεία: -"); self.ch_files.setObjectName("chip")
        self.ch_age = QLabel("Ηλικία: -"); self.ch_age.setObjectName("chip")

        chips_row.addWidget(self.ch_id)
        chips_row.addWidget(self.ch_balance)
        chips_row.addWidget(self.ch_files)
        chips_row.addWidget(self.ch_age)
        chips_row.addStretch(1)
        root.addLayout(chips_row)

        # Progress bar for customer status
        self.progress_bar_preview = QProgressBar()
        self.progress_bar_preview.setTextVisible(True)
        self.progress_bar_preview.setRange(0, 100)
        self.progress_bar_preview.setFixedHeight(20)
        root.addWidget(self.progress_bar_preview)

        # Status chips
        status_row = QHBoxLayout()
        status_row.setSpacing(6)
        self.ch_status = QLabel("Κατάσταση: -"); self.ch_status.setObjectName("chip")
        self.ch_declaration = QLabel("Δήλωση: -"); self.ch_declaration.setObjectName("chip")
        self.ch_amount = QLabel("Ποσό: -"); self.ch_amount.setObjectName("chip")

        status_row.addWidget(self.ch_status)
        status_row.addWidget(self.ch_declaration)
        status_row.addWidget(self.ch_amount)
        status_row.addStretch(1)
        root.addLayout(status_row)

        info = QFrame()
        info.setObjectName("subcard")
        g = QGridLayout(info)
        g.setContentsMargins(10, 10, 10, 10)
        g.setHorizontalSpacing(8)
        g.setVerticalSpacing(6)

        self.v_name = QLabel("-")
        self.v_phone = QLabel("-")
        self.v_afm = QLabel("-")
        self.v_amka = QLabel("-")
        self.v_ama = QLabel("-")
        self.v_service = QLabel("-")
        self.v_date = QLabel("-")
        self.v_payment = QLabel("-")

        for w in (self.v_name, self.v_phone, self.v_afm, self.v_amka, self.v_ama, self.v_service, self.v_date, self.v_payment):
            w.setTextInteractionFlags(Qt.TextSelectableByMouse)
            w.setWordWrap(True)

        g.addWidget(QLabel("Ονοματεπώνυμο:"), 0, 0); g.addWidget(self.v_name, 0, 1)
        g.addWidget(QLabel("Τηλέφωνο:"),      1, 0); g.addWidget(self.v_phone, 1, 1)
        g.addWidget(QLabel("ΑΦΜ:"),           2, 0); g.addWidget(self.v_afm, 2, 1)
        g.addWidget(QLabel("ΑΜΚΑ:"),          3, 0); g.addWidget(self.v_amka, 3, 1)
        g.addWidget(QLabel("ΑΜΑ:"),           4, 0); g.addWidget(self.v_ama, 4, 1)
        g.addWidget(QLabel("Υπηρεσία:"),      5, 0); g.addWidget(self.v_service, 5, 1)
        g.addWidget(QLabel("Ημερομηνία:"),    6, 0); g.addWidget(self.v_date, 6, 1)
        g.addWidget(QLabel("Πληρωμή:"),       7, 0); g.addWidget(self.v_payment, 7, 1)

        root.addWidget(info)

        self.lbl_actions = QLabel("Ενέργειες σήμερα: -")
        self.lbl_actions.setObjectName("muted")
        root.addWidget(self.lbl_actions)

        root.addWidget(QLabel("Αίτημα / Σημειώσεις (προεπισκόπηση):"))
        self.notes_preview = QTextEdit()
        self.notes_preview.setReadOnly(True)
        self.notes_preview.setMinimumHeight(70)
        self.notes_preview.setMaximumHeight(100)
        root.addWidget(self.notes_preview)

        files_box = QFrame()
        files_box.setObjectName("subcard")
        vb = QVBoxLayout(files_box)
        vb.setContentsMargins(10, 10, 10, 10)
        vb.setSpacing(6)

        top = QHBoxLayout()
        self.lbl_files_title = QLabel("Αρχεία πελάτη")
        self.lbl_files_title.setObjectName("subtitle")
        self.lbl_files_hint = QLabel("Επίλεξε → Άνοιγμα / Εκτύπωση")
        self.lbl_files_hint.setObjectName("muted")
        top.addWidget(self.lbl_files_title)
        top.addStretch(1)
        top.addWidget(self.lbl_files_hint)
        vb.addLayout(top)

        self.list_files = QListWidget()
        polish_list(self.list_files)
        self.list_files.setMinimumHeight(100)
        self.list_files.setMaximumHeight(150)
        vb.addWidget(self.list_files)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        self.btn_open_folder = QPushButton("Άνοιγμα φακέλου")
        self.btn_open_folder.setObjectName("secondary")
        self.btn_open_file = QPushButton("Άνοιγμα επιλεγμένου")
        self.btn_open_file.setObjectName("secondary")
        self.btn_print = QPushButton("Εκτύπωση επιλεγμένου")

        btn_row.addWidget(self.btn_open_folder)
        btn_row.addWidget(self.btn_open_file)
        btn_row.addWidget(self.btn_print)
        btn_row.addStretch(1)
        vb.addLayout(btn_row)

        root.addWidget(files_box)

        self._on_open_folder = None
        self._on_open_selected = None
        self._on_print_selected = None

        self.btn_open_folder.clicked.connect(lambda: self._on_open_folder() if self._on_open_folder else None)
        self.btn_open_file.clicked.connect(lambda: self._on_open_selected() if self._on_open_selected else None)
        self.btn_print.clicked.connect(lambda: self._on_print_selected() if self._on_print_selected else None)

    def bind_actions(self, on_open_folder, on_open_selected, on_print_selected):
        self._on_open_folder = on_open_folder
        self._on_open_selected = on_open_selected
        self._on_print_selected = on_print_selected

    def clear(self):
        self._excel_row_index = None
        self._customer_id = None
        self._folder_path = None

        self.lbl_title.setText("Επέλεξε πελάτη")
        self.lbl_sub.setText("Κάνε κλικ σε γραμμή στον πίνακα αριστερά για προεπισκόπηση.")
        self.ch_id.setText("ID: -")
        self.ch_balance.setText("Υπόλοιπο: -")
        self.ch_files.setText("Αρχεία: -")
        self.ch_age.setText("Ηλικία: -")
        self.ch_status.setText("Κατάσταση: -")
        self.ch_declaration.setText("Δήλωση: -")
        self.ch_amount.setText("Ποσό: -")
        self.progress_bar_preview.setValue(0)

        self.v_name.setText("-")
        self.v_phone.setText("-")
        self.v_afm.setText("-")
        self.v_amka.setText("-")
        self.v_ama.setText("-")
        self.v_service.setText("-")
        self.v_date.setText("-")
        self.v_payment.setText("-")

        self.lbl_actions.setText("Ενέργειες σήμερα: -")
        self.notes_preview.setPlainText("")
        self.list_files.clear()

    def set_record(self, excel_row_index: int, rec: dict):
        self._excel_row_index = excel_row_index
        try:
            cid = int(rec.get("ID") or 0)
        except Exception:
            cid = 0
        self._customer_id = cid

        name = str(rec.get("Name") or "-")
        phone = str(rec.get("Phone") or "-")
        afm = str(rec.get("AFM") or "-")
        amka = str(rec.get("AMKA") or "-")
        ama = str(rec.get("AMA") or "-")
        service = str(rec.get("ServiceType") or "-")
        d = str(rec.get("Date") or "-")
        payment = str(rec.get("PaymentMethod") or "-")
        bal = str(rec.get("Balance") or "0")
        created = str(rec.get("CreatedAt") or "")
        
        # Νέα πεδία
        customer_status = str(rec.get("CustomerStatus") or "Νέος")
        declaration_status = str(rec.get("DeclarationStatus") or "-")
        amount = str(rec.get("Amount") or "-")
        goal = str(rec.get("Goal") or "")

        self.lbl_title.setText(f"Πελάτης #{cid}")
        self.lbl_sub.setText(name)

        self.ch_id.setText(f"ID: {cid}")
        self.ch_balance.setText(f"Υπόλοιπο: {bal}")
        self.ch_age.setText(f"Ηλικία: {human_duration(created)}")
        self.ch_status.setText(f"Κατάσταση: {customer_status}")
        self.ch_declaration.setText(f"Δήλωση: {declaration_status}")
        self.ch_amount.setText(f"Ποσό: {amount}")

        # Ορισμός progress bar βάσει status
        status_map = {
            "Νέος": (10, "#7A2B2B"),
            "Σε επεξεργασία": (30, "#7F67BE"),
            "Αναμονή": (50, "#7F67BE"),
            "Ολοκληρωμένος": (100, "#2E6E3F"),
            "Απορριφθείς": (0, "#7A2B2B")
        }
        
        progress, color = status_map.get(customer_status, (0, "#7A2B2B"))
        self.progress_bar_preview.setValue(progress)
        self.progress_bar_preview.setStyleSheet(f"QProgressBar::chunk {{ background-color: {color}; }}")

        folder = resolve_customer_folder_from_record(rec)
        self._folder_path = folder

        files_yes = "Όχι"
        self.list_files.clear()
        if self._folder_path and self._folder_path.exists():
            file_items = [p for p in sorted(self._folder_path.iterdir()) if p.is_file()]
            for p in file_items:
                self.list_files.addItem(QListWidgetItem(p.name))
            files_yes = "Ναι" if file_items else "Όχι"

        self.ch_files.setText(f"Αρχεία: {files_yes}")
        self.v_name.setText(name)
        self.v_phone.setText(phone)
        self.v_afm.setText(afm)
        self.v_amka.setText(amka)
        self.v_ama.setText(ama)
        self.v_service.setText(service)
        self.v_date.setText(d)
        self.v_payment.setText(payment)

        self.lbl_actions.setText(f"Ενέργειες σήμερα: {str(rec.get('ActionsToday') or '-')}")
        
        # Προσθήκη goal στις σημειώσεις αν υπάρχει
        notes = str(rec.get("RequestNotes") or "")
        if goal and goal.strip():
            notes = f"Στόχος: {goal}\n\n{notes}"
        self.notes_preview.setPlainText(notes)

    def selected_file_path(self) -> Optional[Path]:
        item = self.list_files.currentItem()
        if not item or not self._folder_path:
            return None
        p = self._folder_path / item.text()
        return p if p.exists() else None

    @property
    def excel_row_index(self) -> Optional[int]:
        return self._excel_row_index

    @property
    def customer_id(self) -> Optional[int]:
        return self._customer_id

    @property
    def folder_path(self) -> Optional[Path]:
        return self._folder_path


# ----------------------------- Login -----------------------------
class LoginWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Client Manager v1 – Σύνδεση")
        self.resize(440, 260)

        root = QVBoxLayout(self)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(12)

        title = QLabel("Σύνδεση")
        title.setObjectName("title")
        root.addWidget(title)

        hint = QLabel("Προεπιλογή: admin / 1234 (άλλαξε το μετά τη σύνδεση)")
        hint.setObjectName("muted")
        root.addWidget(hint)

        card = QFrame()
        card.setObjectName("card")
        c = QVBoxLayout(card)
        c.setContentsMargins(14, 14, 14, 14)
        c.setSpacing(10)

        self.in_user = QLineEdit()
        self.in_user.setPlaceholderText("Όνομα χρήστη")

        self.in_pass = QLineEdit()
        self.in_pass.setPlaceholderText("Κωδικός")
        self.in_pass.setEchoMode(QLineEdit.Password)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        self.btn_login = QPushButton("Σύνδεση")
        self.btn_login.clicked.connect(self.do_login)

        self.btn_change = QPushButton("Αλλαγή κωδικού admin")
        self.btn_change.setObjectName("secondary")
        self.btn_change.clicked.connect(self.change_admin_password)

        btn_row.addWidget(self.btn_login)
        btn_row.addWidget(self.btn_change)

        c.addWidget(self.in_user)
        c.addWidget(self.in_pass)
        c.addLayout(btn_row)

        root.addWidget(card)

    def do_login(self):
        try:
            u = self.in_user.text().strip()
            p = self.in_pass.text()
            if not u or not p:
                QMessageBox.warning(self, "Λείπει κάτι", "Βάλε όνομα χρήστη και κωδικό.")
                return
            if verify_user(u, p):
                audit_log(u, "LOGIN")
                self.main = MainWindow(current_user=u)
                self.main.show()
                self.close()
            else:
                QMessageBox.warning(self, "Αποτυχία σύνδεσης", "Λάθος στοιχεία.")
        except Exception as e:
            QMessageBox.critical(self, "Σφάλμα", f"{e}\n\n{traceback.format_exc()}")

    def change_admin_password(self):
        dlg = ScrollableDialog(self)
        dlg.setWindowTitle("Αλλαγή κωδικού admin")
        dlg.resize(420, 220)

        p1 = QLineEdit(); p1.setPlaceholderText("Νέος κωδικός"); p1.setEchoMode(QLineEdit.Password)
        p2 = QLineEdit(); p2.setPlaceholderText("Επανάληψη νέου κωδικού"); p2.setEchoMode(QLineEdit.Password)

        btn = QPushButton("Αποθήκευση")

        def save():
            if not p1.text() or p1.text() != p2.text():
                QMessageBox.warning(dlg, "Σφάλμα", "Οι κωδικοί δεν ταιριάζουν.")
                return
            set_user_password("admin", p1.text())
            QMessageBox.information(dlg, "ΟΚ", "Ο κωδικός του admin ενημερώθηκε.")
            dlg.accept()

        btn.clicked.connect(save)

        dlg.addWidget(QLabel("Θέσε νέο κωδικό για τον χρήστη 'admin'."))
        dlg.addWidget(p1); dlg.addWidget(p2); dlg.addWidget(btn)
        dlg.exec()


# ----------------------------- Enhanced Customer Dialog -----------------------------
class CustomerDialog(ScrollableDialog):
    def __init__(self, current_user: str, excel_row_index: int):
        super().__init__()
        self.current_user = current_user
        self.row_index = excel_row_index
        self.setWindowTitle("Καρτέλα πελάτη")
        self.resize(1200, 850)  # Αύξηση ύψους για τα νέα στοιχεία

        wb, ws = open_ws()
        self.wb = wb
        self.ws = ws
        self.rec = row_to_record(ws, excel_row_index)
        self.customer_id = int(self.rec.get("ID") or 0)

        title = QLabel(f"Πελάτης #{self.customer_id}")
        title.setObjectName("title")
        self.addWidget(title)

        subtitle = QLabel(str(self.rec.get("Name") or ""))
        subtitle.setObjectName("muted")
        self.addWidget(subtitle)

        # Status Bar Section
        status_section = QFrame()
        status_section.setObjectName("card")
        status_layout = QVBoxLayout(status_section)
        status_layout.setContentsMargins(12, 12, 12, 12)
        status_layout.setSpacing(8)
        
        # Progress Bar για Customer Status
        self.progress_label = QLabel("Κατάσταση Πελάτη:")
        self.progress_label.setObjectName("subtitle")
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setRange(0, 100)
        
        # Customer Status ComboBox
        self.customer_status_combo = QComboBox()
        self.customer_status_combo.addItems([
            "Νέος",
            "Σε επεξεργασία", 
            "Αναμονή",
            "Ολοκληρωμένος",
            "Απορριφθείς"
        ])
        self.customer_status_combo.currentTextChanged.connect(self.on_customer_status_changed)
        
        status_row = QHBoxLayout()
        status_row.addWidget(QLabel("Κατάσταση:"))
        status_row.addWidget(self.customer_status_combo)
        status_row.addStretch(1)
        status_row.addWidget(QLabel("Πρόοδος:"))
        status_row.addWidget(self.progress_bar)
        status_row.setSpacing(10)
        
        status_layout.addLayout(status_row)
        
        # Goal Section
        goal_group = QFrame()
        goal_group.setObjectName("subcard")
        goal_layout = QVBoxLayout(goal_group)
        goal_layout.setContentsMargins(10, 10, 10, 10)
        
        self.goal_label = QLabel("Στόχος Πελάτη:")
        self.goal_label.setObjectName("subtitle")
        
        self.goal_text = QTextEdit()
        self.goal_text.setPlaceholderText("Περιγράψτε τον στόχο του πελάτη...")
        self.goal_text.setMaximumHeight(80)
        
        goal_layout.addWidget(self.goal_label)
        goal_layout.addWidget(self.goal_text)
        
        status_layout.addWidget(goal_group)
        
        # Declaration Status
        decl_group = QFrame()
        decl_group.setObjectName("subcard")
        decl_layout = QVBoxLayout(decl_group)
        decl_layout.setContentsMargins(10, 10, 10, 10)
        
        self.declaration_label = QLabel("Κατάσταση Δηλώσεων:")
        self.declaration_label.setObjectName("subtitle")
        
        self.declaration_combo = QComboBox()
        self.declaration_combo.addItems([
            "Μη Ορισμένη",
            "Υποβολή Ε1",
            "Υποβολή Ε2", 
            "Υποβολή Ε3",
            "Αναμονή Απάντησης",
            "Ολοκληρωμένη",
            "Πρόβλημα"
        ])
        
        decl_layout.addWidget(self.declaration_label)
        decl_layout.addWidget(self.declaration_combo)
        
        status_layout.addWidget(decl_group)
        
        self.addWidget(status_section)

        chips = QHBoxLayout()
        chips.setSpacing(6)
        self.ch_phone = QLabel(f"Τηλέφωνο: {self.rec.get('Phone')}")
        self.ch_phone.setObjectName("chip")
        self.ch_balance = QLabel(f"Υπόλοιπο: {self.rec.get('Balance')}")
        self.ch_balance.setObjectName("chip")
        self.ch_folder = QLabel("Φάκελος: " + (str(self.rec.get("FolderPath") or "-")))
        self.ch_folder.setObjectName("chip")
        chips.addWidget(self.ch_phone)
        chips.addWidget(self.ch_balance)
        chips.addWidget(self.ch_folder)
        chips.addStretch(1)
        self.addLayout(chips)

        extra_card = QFrame()
        extra_card.setObjectName("card")
        extra_l = QGridLayout(extra_card)
        extra_l.setContentsMargins(12, 12, 12, 12)
        extra_l.setHorizontalSpacing(10)
        extra_l.setVerticalSpacing(8)

        self.ed_afm = QLineEdit(str(self.rec.get("AFM") or ""))
        self.ed_amka = QLineEdit(str(self.rec.get("AMKA") or ""))
        self.ed_ama = QLineEdit(str(self.rec.get("AMA") or ""))
        
        # Νέο πεδίο για Amount
        self.ed_amount = QLineEdit(str(self.rec.get("Amount") or ""))
        self.ed_amount.setPlaceholderText("Ποσό (π.χ. 150.00)")
        self.ed_amount.textChanged.connect(self.on_amount_changed)
        
        # Νέο πεδίο για AMKA Valid
        self.cb_amka_valid = QComboBox()
        self.cb_amka_valid.addItems(["Άγνωστο", "Έγκυρο", "Μη Έγκυρο"])

        set_digits_only(self.ed_afm, 9)
        set_digits_only(self.ed_amka, 11)  # Μόνο 11 ψηφία, όχι ημερομηνία
        self.ed_ama.setMaxLength(20)

        extra_l.addWidget(QLabel("ΑΦΜ:"), 0, 0); extra_l.addWidget(self.ed_afm, 0, 1)
        extra_l.addWidget(QLabel("ΑΜΚΑ:"), 1, 0); extra_l.addWidget(self.ed_amka, 1, 1)
        extra_l.addWidget(QLabel("ΑΜΑ:"),  2, 0); extra_l.addWidget(self.ed_ama,  2, 1)
        extra_l.addWidget(QLabel("Ποσό:"), 3, 0); extra_l.addWidget(self.ed_amount, 3, 1)
        extra_l.addWidget(QLabel("AMKA Έγκυρο:"), 4, 0); extra_l.addWidget(self.cb_amka_valid, 4, 1)

        self.addWidget(QLabel("Στοιχεία"))
        self.addWidget(extra_card)

        self.addWidget(QLabel("Αίτημα / Σημειώσεις"))
        self.notes = QTextEdit()
        self.notes.setText(str(self.rec.get("RequestNotes") or ""))
        self.notes.setMinimumHeight(80)
        self.addWidget(self.notes)

        files_card = QFrame()
        files_card.setObjectName("card")
        files_layout = QVBoxLayout(files_card)
        files_layout.setContentsMargins(12, 12, 12, 12)
        files_layout.setSpacing(8)

        files_layout.addWidget(QLabel("Αρχεία πελάτη"))

        self.list_files = QListWidget()
        polish_list(self.list_files)
        self.list_files.setMinimumHeight(120)
        files_layout.addWidget(self.list_files)

        files_btn_row = QHBoxLayout()
        files_btn_row.setSpacing(6)
        self.btn_open_folder = QPushButton("Άνοιγμα φακέλου")
        self.btn_open_folder.setObjectName("secondary")
        self.btn_open_folder.clicked.connect(self.open_folder)

        self.btn_open_file = QPushButton("Άνοιγμα επιλεγμένου")
        self.btn_open_file.setObjectName("secondary")
        self.btn_open_file.clicked.connect(self.open_selected_file)

        self.btn_print = QPushButton("Εκτύπωση επιλεγμένου")
        self.btn_print.clicked.connect(self.print_selected_file)

        files_btn_row.addWidget(self.btn_open_folder)
        files_btn_row.addWidget(self.btn_open_file)
        files_btn_row.addWidget(self.btn_print)
        files_btn_row.addStretch(1)
        files_layout.addLayout(files_btn_row)

        self.addWidget(files_card)

        audit_card = QFrame()
        audit_card.setObjectName("card")
        audit_layout = QVBoxLayout(audit_card)
        audit_layout.setContentsMargins(12, 12, 12, 12)
        audit_layout.setSpacing(8)

        audit_layout.addWidget(QLabel("Ιστορικό (audit log)"))
        self.audit_box = QTextEdit()
        self.audit_box.setReadOnly(True)
        self.audit_box.setMinimumHeight(150)
        audit_layout.addWidget(self.audit_box)

        self.addWidget(audit_card)

        bottom = QHBoxLayout()
        bottom.setSpacing(8)
        self.btn_refresh = QPushButton("Ανανέωση")
        self.btn_refresh.setObjectName("secondary")
        self.btn_refresh.clicked.connect(self.refresh_views)

        self.btn_save = QPushButton("Αποθήκευση αλλαγών")
        self.btn_save.clicked.connect(self.save_changes)

        bottom.addWidget(self.btn_save)
        bottom.addWidget(self.btn_refresh)
        bottom.addStretch(1)
        self.addLayout(bottom)

        self.refresh_views()

    def on_customer_status_changed(self, status: str):
        """Ενημερώνει τη μπάρα προόδου βάσει της κατάστασης"""
        status_map = {
            "Νέος": 10,
            "Σε επεξεργασία": 30,
            "Αναμονή": 50,
            "Ολοκληρωμένος": 100,
            "Απορριφθείς": 0
        }
        progress = status_map.get(status, 0)
        self.progress_bar.setValue(progress)
        
        # Αλλαγή χρώματος βάσει προόδου
        if progress >= 70:
            self.progress_bar.setStyleSheet("QProgressBar::chunk { background-color: #2E6E3F; }")
        elif progress >= 30:
            self.progress_bar.setStyleSheet("QProgressBar::chunk { background-color: #7F67BE; }")
        else:
            self.progress_bar.setStyleSheet("QProgressBar::chunk { background-color: #7A2B2B; }")
    
    def on_amount_changed(self, text: str):
        """Επικύρωση του ποσού"""
        try:
            # Αφαίρεση όλων των χαρακτήρων εκτός από ψηφία και τελεία
            clean = ''.join(c for c in text if c.isdigit() or c == '.')
            if clean.count('.') > 1:
                # Άφησε μόνο την πρώτη τελεία
                parts = clean.split('.')
                clean = parts[0] + '.' + ''.join(parts[1:])
            
            if clean != text:
                self.ed_amount.setText(clean)
                self.ed_amount.setCursorPosition(len(clean))
        except:
            pass

    def refresh_views(self):
        self.rec = row_to_record(self.ws, self.row_index)

        self.ch_phone.setText(f"Τηλέφωνο: {self.rec.get('Phone')}")
        self.ch_balance.setText(f"Υπόλοιπο: {self.rec.get('Balance')}")
        
        folder = resolve_customer_folder_from_record(self.rec)
        if folder:
            # εμφάνιση στο chip (να βλέπεις το πραγματικό path)
            self.ch_folder.setText("Φάκελος: " + str(folder))

            self.list_files.clear()
            for p in sorted(folder.iterdir()):
                if p.is_file():
                    self.list_files.addItem(QListWidgetItem(p.name))
        else:
            self.list_files.clear()

        lines = read_audit_for_customer(self.customer_id)
        if not lines:
            self.audit_box.setPlainText("(δεν υπάρχει ιστορικό)")
        else:
            txt = []
            for ts, user, action, cid, details in lines[-300:]:
                txt.append(f"{ts} | {user} | {action} | {details}")
            self.audit_box.setPlainText("\n".join(txt))

        # Νέα πεδία
        goal = str(self.rec.get("Goal") or "")
        declaration_status = str(self.rec.get("DeclarationStatus") or "Μη Ορισμένη")
        customer_status = str(self.rec.get("CustomerStatus") or "Νέος")
        amount = str(self.rec.get("Amount") or "")
        amka_valid = str(self.rec.get("AMKA_Valid") or "Άγνωστο")
        
        # Ορισμός τιμών
        self.goal_text.setPlainText(goal)
        
        # Βρίσκουμε το index για το declaration combo
        dec_index = self.declaration_combo.findText(declaration_status)
        if dec_index >= 0:
            self.declaration_combo.setCurrentIndex(dec_index)
        
        # Βρίσκουμε το index για το customer status combo
        status_index = self.customer_status_combo.findText(customer_status)
        if status_index >= 0:
            self.customer_status_combo.setCurrentIndex(status_index)
        else:
            self.customer_status_combo.setCurrentText("Νέος")
        
        # Ορισμός ποσού
        if amount:
            try:
                # Μορφοποίηση με 2 δεκαδικά
                amount_float = float(amount)
                self.ed_amount.setText(f"{amount_float:.2f}")
            except:
                self.ed_amount.setText("")
        
        # Ορισμός AMKA Valid
        valid_index = self.cb_amka_valid.findText(amka_valid)
        if valid_index >= 0:
            self.cb_amka_valid.setCurrentIndex(valid_index)
        else:
            self.cb_amka_valid.setCurrentText("Άγνωστο")

        # sync fields with latest record (ώστε να μην βλέπεις "παλιά/άκυρα")
        self.ed_afm.setText(str(self.rec.get("AFM") or ""))
        self.ed_amka.setText(str(self.rec.get("AMKA") or ""))
        self.ed_ama.setText(str(self.rec.get("AMA") or ""))
        self.notes.setPlainText(str(self.rec.get("RequestNotes") or ""))

    def open_folder(self):
        folder = resolve_customer_folder_from_record(self.rec)
        if not folder:
            QMessageBox.information(self, "Πληροφορία", "Δεν υπάρχει αποθηκευμένος/εντοπίσιμος φάκελος για τον πελάτη.")
            return
        if not folder.exists():
            QMessageBox.warning(self, "Σφάλμα", "Ο φάκελος δεν υπάρχει στο δίσκο.")
            return
        open_in_explorer(folder)

    def selected_file_path(self) -> Optional[Path]:
        folder = resolve_customer_folder_from_record(self.rec)
        if not folder:
            return None
        item = self.list_files.currentItem()
        if not item:
            return None
        p = folder / item.text()
        return p if p.exists() else None

    def open_selected_file(self):
        p = self.selected_file_path()
        if not p:
            return
        open_in_explorer(p)

    def print_selected_file(self):
        p = self.selected_file_path()
        if not p:
            return
        ok = print_with_dialog(p, self)
        if not ok:
            QMessageBox.information(self, "Εκτύπωση", "Δεν μπόρεσα να κάνω άμεση εκτύπωση. Θα ανοίξω το αρχείο.")
            open_in_explorer(p)
        audit_log(self.current_user, "PRINT", self.customer_id, f"{p.name}")

    def save_changes(self):
        new_notes = self.notes.toPlainText()
        new_afm = (self.ed_afm.text() or "").strip()
        new_amka = (self.ed_amka.text() or "").strip()
        new_ama = (self.ed_ama.text() or "").strip()
        
        # Νέα πεδία
        new_goal = self.goal_text.toPlainText()
        new_declaration = self.declaration_combo.currentText()
        new_customer_status = self.customer_status_combo.currentText()
        
        # Επεξεργασία ποσού
        amount_text = (self.ed_amount.text() or "").strip().replace(",", ".")
        new_amount = ""
        if amount_text:
            try:
                amount_float = float(amount_text)
                new_amount = f"{amount_float:.2f}"
            except:
                new_amount = ""
        
        new_amka_valid = self.cb_amka_valid.currentText()

        if new_afm and len(new_afm) != 9:
            QMessageBox.warning(self, "Σφάλμα", "Το ΑΦΜ πρέπει να έχει 9 ψηφία (ή να μείνει κενό).")
            return
        if new_amka and len(new_amka) != 11:
            QMessageBox.warning(self, "Σφάλμα", "Το ΑΜΚΑ πρέπει να έχει 11 ψηφία (ή να μείνει κενό).")
            return

        update_row(self.ws, self.row_index, {
            "AFM": new_afm,
            "AMKA": new_amka,
            "AMA": new_ama,
            "RequestNotes": new_notes,
            "Goal": new_goal,
            "DeclarationStatus": new_declaration,
            "CustomerStatus": new_customer_status,
            "Amount": new_amount,
            "AMKA_Valid": new_amka_valid,
            "LastEditedBy": self.current_user,
            "LastEditedAt": now_iso(),
        })
        self.wb.save(EXCEL_PATH)
        audit_log(self.current_user, "EDIT", self.customer_id, 
                 f"Ενημέρωση: Goal, Status={new_customer_status}, Amount={new_amount}")
        QMessageBox.information(self, "ΟΚ", "Ο πελάτης ενημερώθηκε.")
        self.refresh_views()


# ----------------------------- Main Window -----------------------------
class MainWindow(QWidget):
    def __init__(self, current_user: str):
        super().__init__()
        self.user = current_user
        self.setWindowTitle("Client Manager v1")

        self.setMinimumSize(900, 600)
        self.resize(1200, 750)

        self.pending_files: List[str] = []
        self.last_saved_customer_id: Optional[int] = None
        self.last_saved_customer_folder: Optional[Path] = None
        self.search_selected_excel_row: Optional[int] = None

        root = QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        # Header
        header = QHBoxLayout()
        title = QLabel("Client Manager")
        title.setObjectName("title")

        self.lbl_status = QLabel(f"Συνδεδεμένος: {self.user}")
        self.lbl_status.setObjectName("muted")
        self.lbl_status.setWordWrap(True)

        header.addWidget(title)
        header.addStretch(1)

        self.theme_switch = QCheckBox("")
        self.theme_switch.setObjectName("themeSwitch")
        self.theme_switch.setToolTip("Εναλλαγή Dark/Light")
        self.theme_switch.stateChanged.connect(self.on_toggle_theme)

        self.lbl_theme = QLabel("Theme")
        self.lbl_theme.setObjectName("muted")

        header.addWidget(self.lbl_theme)
        header.addWidget(self.theme_switch)
        header.addSpacing(10)
        header.addWidget(self.lbl_status)
        root.addLayout(header)

        # Tabs
        self.tabs = QTabWidget()
        root.addWidget(self.tabs)

        self.tab_new = ResponsiveScrollArea()
        self.tab_new_inner = QWidget()
        self.tab_new.setWidget(self.tab_new_inner)

        self.tab_search = ResponsiveScrollArea()
        self.tab_search_inner = QWidget()
        self.tab_search.setWidget(self.tab_search_inner)

        self.tab_dashboard = ResponsiveScrollArea()
        self.tab_dashboard_inner = QWidget()
        self.tab_dashboard.setWidget(self.tab_dashboard_inner)

        self.tabs.addTab(self.tab_new, "Νέος πελάτης")
        self.tabs.addTab(self.tab_search, "Αναζήτηση")
        self.tabs.addTab(self.tab_dashboard, "Πίνακας")
        self.tabs_settings = ResponsiveScrollArea()
        self.tab_settings_inner = QWidget()
        self.tabs_settings.setWidget(self.tab_settings_inner)
        self.tabs.addTab(self.tabs_settings, "Ρυθμίσεις")
        

        self.build_tab_new(self.tab_new_inner)
        self.build_tab_search(self.tab_search_inner)
        self.build_tab_dashboard(self.tab_dashboard_inner)
        self.build_tab_settings(self.tab_settings_inner)

        self.apply_theme(get_theme())
        self.refresh_next_id()
        self.refresh_dashboard()

    # ---------------- Theme ----------------
    def apply_theme(self, theme: str):
        app = QApplication.instance()
        if not app:
            return
        if theme == "light":
            app.setStyleSheet(LIGHT_STYLE)
            self.theme_switch.blockSignals(True)
            self.theme_switch.setChecked(True)
            self.theme_switch.blockSignals(False)
            self.lbl_theme.setText("Light")
        else:
            app.setStyleSheet(DARK_STYLE)
            self.theme_switch.blockSignals(True)
            self.theme_switch.setChecked(False)
            self.theme_switch.blockSignals(False)
            self.lbl_theme.setText("Dark")
        set_theme(theme)

    def on_toggle_theme(self):
        self.apply_theme("light" if self.theme_switch.isChecked() else "dark")

    # ---------------- Tab: Νέος Πελάτης ----------------
    def build_tab_new(self, parent: QWidget):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setSpacing(10)

        # Settings card
        settings_card = QFrame()
        settings_card.setObjectName("card")
        sc = QVBoxLayout(settings_card)
        sc.setContentsMargins(12, 12, 12, 12)
        sc.setSpacing(8)

        sc.addWidget(QLabel("Ρυθμίσεις αποθήκευσης αρχείων"))

        row_root = QHBoxLayout()
        row_root.setSpacing(8)
        self.in_clients_root = QLineEdit(str(get_clients_root()))
        self.in_clients_root.setPlaceholderText("Ρίζα φακέλων πελατών (π.χ. D:\\Clients)")

        btn_browse_root = QPushButton("Επιλογή…")
        btn_browse_root.setObjectName("secondary")
        btn_browse_root.clicked.connect(self.browse_clients_root)

        btn_save_root = QPushButton("Αποθήκευση")
        btn_save_root.setObjectName("secondary")
        btn_save_root.clicked.connect(self.save_clients_root)

        row_root.addWidget(self.in_clients_root, 4)
        row_root.addWidget(btn_browse_root, 1)
        row_root.addWidget(btn_save_root, 1)
        sc.addLayout(row_root)

        hint = QLabel("Σημαντικό: Αν αλλάξεις ρίζα μετά, οι παλιοί πελάτες μπορεί να δείχνουν σε παλιό FolderPath.")
        hint.setObjectName("musted")
        hint.setWordWrap(True)
        sc.addWidget(hint)

        lay.addWidget(settings_card)

        # Form card
        form_card = QFrame()
        form_card.setObjectName("card")
        c = QVBoxLayout(form_card)
        c.setContentsMargins(12, 12, 12, 12)
        c.setSpacing(8)

        top = QHBoxLayout()
        top.setSpacing(8)
        self.service_type = QComboBox()
        self.service_type.addItems(["ΚΕΠ", "Νομικά"])

        self.lbl_next_id = QLabel("Επόμενο ID: -")
        self.lbl_next_id.setObjectName("chip")

        self.lbl_date = QLabel(date.today().isoformat())
        self.lbl_date.setObjectName("chip")

        top.addWidget(QLabel("Τύπος υπηρεσίας:"))
        top.addWidget(self.service_type)
        top.addStretch(1)
        top.addWidget(QLabel("Ημερομηνία:"))
        top.addWidget(self.lbl_date)
        top.addSpacing(8)
        top.addWidget(self.lbl_next_id)
        c.addLayout(top)

        grid = QGridLayout()
        grid.setHorizontalSpacing(10)
        grid.setVerticalSpacing(8)

        self.in_name = QLineEdit(); self.in_name.setPlaceholderText("Ονοματεπώνυμο")
        self.in_phone = QLineEdit(); self.in_phone.setPlaceholderText("Τηλέφωνο")
        self.in_afm = QLineEdit(); self.in_afm.setPlaceholderText("ΑΦΜ (9 ψηφία)")
        self.in_amka = QLineEdit(); self.in_amka.setPlaceholderText("ΑΜΚΑ (11 ψηφία)")
        self.in_ama = QLineEdit(); self.in_ama.setPlaceholderText("ΑΜΑ (προαιρετικό)")

        set_phone_validator(self.in_phone, 15)
        set_digits_only(self.in_afm, 9)
        # AMKA: Μόνο 11 ψηφία, όχι ημερομηνία
        self.in_amka.setMaxLength(11)
        rx_amka = QRegularExpression(r"^\d{0,11}$")
        self.in_amka.setValidator(QRegularExpressionValidator(rx_amka))
        self.in_ama.setMaxLength(20)

        grid.addWidget(QLabel("Ονοματεπώνυμο:"), 0, 0)
        grid.addWidget(self.in_name, 0, 1, 1, 3)
        grid.addWidget(QLabel("Τηλέφωνο:"), 1, 0)
        grid.addWidget(self.in_phone, 1, 1)
        grid.addWidget(QLabel("ΑΦΜ:"), 1, 2)
        grid.addWidget(self.in_afm, 1, 3)
        grid.addWidget(QLabel("ΑΜΚΑ:"), 2, 0)
        grid.addWidget(self.in_amka, 2, 1)
        grid.addWidget(QLabel("ΑΜΑ:"), 2, 2)
        grid.addWidget(self.in_ama, 2, 3)

        c.addLayout(grid)

        row_tax = QHBoxLayout()
        row_tax.setSpacing(8)
        self.cb_has_taxis = QComboBox()
        self.cb_has_taxis.addItems(["Όχι", "Ναι"])
        self.cb_has_taxis.currentTextChanged.connect(self.on_has_taxis_changed)
        row_tax.addWidget(QLabel("Έχει Taxisnet;"))
        row_tax.addWidget(self.cb_has_taxis)
        row_tax.addStretch(1)
        c.addLayout(row_tax)

        self.taxis_box = QFrame()
        self.taxis_box.setObjectName("subcard")
        tb = QGridLayout(self.taxis_box)
        tb.setContentsMargins(10, 10, 10, 10)
        tb.setHorizontalSpacing(8)
        tb.setVerticalSpacing(6)

        self.in_taxis_user = QLineEdit(); self.in_taxis_user.setPlaceholderText("Username Taxisnet")
        self.in_taxis_pass = QLineEdit(); self.in_taxis_pass.setPlaceholderText("Password Taxisnet")
        self.in_taxis_pass.setEchoMode(QLineEdit.Password)
        self.in_taxis_user.setMaxLength(64)
        self.in_taxis_pass.setMaxLength(64)

        tb.addWidget(QLabel("Username:"), 0, 0)
        tb.addWidget(self.in_taxis_user, 0, 1)
        tb.addWidget(QLabel("Password:"), 1, 0)
        tb.addWidget(self.in_taxis_pass, 1, 1)

        self.taxis_box.setVisible(False)
        c.addWidget(self.taxis_box)

        row3 = QHBoxLayout()
        row3.setSpacing(8)
        self.in_kleid = QComboBox(); self.in_kleid.addItems(["Επιβεβαιώθηκε", "Απορρίφθηκε"])
        self.in_amka_status = QComboBox(); self.in_amka_status.addItems(["Επιβεβαιώθηκε", "Απορρίφθηκε"])
        self.in_aporr = QComboBox(); self.in_aporr.addItems(["Όχι", "Ναι"])

        row3.addWidget(QLabel("Κλειδάριθμος:"))
        row3.addWidget(self.in_kleid)
        row3.addSpacing(8)
        row3.addWidget(QLabel("ΑΜΚΑ/ΑΜΑ status:"))
        row3.addWidget(self.in_amka_status)
        row3.addSpacing(8)
        row3.addWidget(QLabel("Απόρριψη:"))
        row3.addWidget(self.in_aporr)
        row3.addStretch(1)
        c.addLayout(row3)

        c.addWidget(QLabel("Επιβεβαίωση πριν την αποθήκευση:"))
        self.cb_confirm_data = QCheckBox("Έχω επιβεβαιώσει τα στοιχεία του πελάτης")
        self.cb_confirm_services = QCheckBox("Έχω επιβεβαιώσει υπηρεσίες & τιμές")
        c.addWidget(self.cb_confirm_data)
        c.addWidget(self.cb_confirm_services)

        c.addWidget(QLabel("Ενέργειες σήμερα: (διάλεξε τουλάχιστον 1)"))
        actions_grid = QGridLayout()
        actions_grid.setHorizontalSpacing(10)
        actions_grid.setVerticalSpacing(6)

        self.service_checks: List[QCheckBox] = []
        for idx, (name, price) in enumerate(SERVICES):
            cb = QCheckBox(f"{name} ({price})")
            cb.stateChanged.connect(self.recalculate)
            self.service_checks.append(cb)
            r = idx // 2
            col = idx % 2
            actions_grid.addWidget(cb, r, col)
        c.addLayout(actions_grid)

        pay = QHBoxLayout()
        pay.setSpacing(8)
        self.cmb_payment = QComboBox()
        self.cmb_payment.addItems(PAYMENT_METHODS)

        self.in_paid = QLineEdit()
        self.in_paid.setPlaceholderText("Πληρώθηκε σήμερα (ποσό)")
        self.in_paid.textChanged.connect(self.recalculate)
        self.in_paid.setFixedWidth(180)
        self.in_paid.setMaxLength(12)

        btn_calc = QPushButton("Υπολογισμός")
        btn_calc.setObjectName("secondary")
        btn_calc.clicked.connect(self.recalculate)

        self.lbl_total = QLabel("Σύνολο: 0.00"); self.lbl_total.setObjectName("chip")
        self.lbl_balance = QLabel("Υπόλοιπο: 0.00"); self.lbl_balance.setObjectName("chip")

        pay.addWidget(QLabel("Πληρωμή:"))
        pay.addWidget(self.cmb_payment)
        pay.addSpacing(8)
        pay.addWidget(self.in_paid)
        pay.addWidget(btn_calc)
        pay.addStretch(1)
        pay.addWidget(self.lbl_total)
        pay.addWidget(self.lbl_balance)
        c.addLayout(pay)

        c.addWidget(QLabel("Αίτημα / Σημειώσεις:"))
        self.in_request = QTextEdit()
        self.in_request.setPlaceholderText("Γράψε το αίτημα του πελάτη / σημειώσεις")
        self.in_request.setMinimumHeight(80)
        self.in_request.setMaximumHeight(120)
        c.addWidget(self.in_request)

        btns = QHBoxLayout()
        btns.setSpacing(8)
        self.btn_save_customer = QPushButton("Αποθήκευση πελάτη στο Excel")
        self.btn_save_customer.clicked.connect(self.save_customer)

        self.btn_open_folder = QPushButton("Άνοιγμα φακέλου τελευταίου πελάτη")
        self.btn_open_folder.setObjectName("secondary")
        self.btn_open_folder.clicked.connect(self.open_last_folder)

        btns.addWidget(self.btn_save_customer, 2)
        btns.addWidget(self.btn_open_folder, 1)
        c.addLayout(btns)

        lay.addWidget(form_card)

        # Files card
        files_card = QFrame()
        files_card.setObjectName("card")
        c2 = QVBoxLayout(files_card)
        c2.setContentsMargins(12, 12, 12, 12)
        c2.setSpacing(8)

        scanner_row = QHBoxLayout()
        scanner_row.setSpacing(8)
        self.in_scanner = QLineEdit(get_scanner_folder())
        self.in_scanner.setPlaceholderText("Φάκελος scanner (προαιρετικό)")

        btn_browse_scanner_folder = QPushButton("Φάκελος…")
        btn_browse_scanner_folder.setObjectName("secondary")
        btn_browse_scanner_folder.clicked.connect(self.browse_scanner_folder)

        btn_browse_scanner_file = QPushButton("Αρχείο…")
        btn_browse_scanner_file.setObjectName("secondary")
        btn_browse_scanner_file.clicked.connect(self.browse_scanner_file)

        btn_set_scanner = QPushButton("Αποθήκευση")
        btn_set_scanner.setObjectName("secondary")
        btn_set_scanner.clicked.connect(self.set_scanner_folder_ui)

        scanner_row.addWidget(self.in_scanner, 4)
        scanner_row.addWidget(btn_browse_scanner_folder, 1)
        scanner_row.addWidget(btn_browse_scanner_file, 1)
        scanner_row.addWidget(btn_set_scanner, 1)
        c2.addLayout(scanner_row)

        c2.addWidget(QLabel("Αρχεία (drag & drop μετά την αποθήκευση πελάτη):"))
        self.drop = DropZone(self.on_files_dropped)
        c2.addWidget(self.drop)

        lists_row = QHBoxLayout()
        lists_row.setSpacing(10)

        pending_col = QVBoxLayout()
        pending_col.setSpacing(4)
        pending_col.addWidget(QLabel("Pending αρχεία"))
        self.list_pending = QListWidget()
        polish_list(self.list_pending)
        self.list_pending.setMinimumHeight(100)
        pending_col.addWidget(self.list_pending)
        lists_row.addLayout(pending_col, 1)

        folder_col = QVBoxLayout()
        folder_col.setSpacing(4)
        folder_col.addWidget(QLabel("Αρχεία φακέλου πελάτη"))
        self.list_folder = QListWidget()
        polish_list(self.list_folder)
        self.list_folder.setMinimumHeight(100)
        folder_col.addWidget(self.list_folder)
        lists_row.addLayout(folder_col, 1)

        scanner_col = QVBoxLayout()
        scanner_col.setSpacing(4)
        scanner_col.addWidget(QLabel("Αρχεία scanner"))
        self.list_scanner = QListWidget()
        polish_list(self.list_scanner)
        self.list_scanner.setMinimumHeight(100)
        scanner_col.addWidget(self.list_scanner)
        lists_row.addLayout(scanner_col, 1)

        c2.addLayout(lists_row)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(6)
        btn_pick = QPushButton("Προσθήκη αρχείων…")
        btn_pick.setObjectName("secondary")
        btn_pick.clicked.connect(self.pick_files)

        btn_move_pending = QPushButton("Μεταφορά pending")
        btn_move_pending.clicked.connect(self.move_pending_to_customer)

        btn_move_scanner = QPushButton("Μεταφορά scanner")
        btn_move_scanner.setObjectName("secondary")
        btn_move_scanner.clicked.connect(self.move_selected_scanner_to_customer)

        btn_refresh_lists = QPushButton("Ανανέωση")
        btn_refresh_lists.setObjectName("secondary")
        btn_refresh_lists.clicked.connect(self.refresh_file_lists)

        btn_print = QPushButton("Εκτύπωση")
        btn_print.clicked.connect(self.print_selected_customer_file)

        btn_confirm_files = QPushButton("Επιβεβαίωση")
        btn_confirm_files.clicked.connect(self.confirm_files)

        btn_row.addWidget(btn_pick)
        btn_row.addWidget(btn_move_pending)
        btn_row.addWidget(btn_move_scanner)
        btn_row.addWidget(btn_print)
        btn_row.addWidget(btn_confirm_files)
        btn_row.addWidget(btn_refresh_lists)
        c2.addLayout(btn_row)

        lay.addWidget(files_card)
        lay.addStretch(1)

    def browse_clients_root(self):
        folder = QFileDialog.getExistingDirectory(self, "Επίλεξε ρίζα φακέλων πελατών")
        if folder:
            self.in_clients_root.setText(folder)

    def save_clients_root(self):
        path = self.in_clients_root.text().strip()
        if not path:
            QMessageBox.warning(self, "Σφάλμα", "Δώσε ένα path για ρίζα φακέλων.")
            return
        p = Path(path)
        try:
            p.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            QMessageBox.warning(self, "Σφάλμα", f"Δεν μπορώ να δημιουργήσω/χρησιμοποιήσω το φάκελο:\n{e}")
            return
        set_clients_root(str(p))
        QMessageBox.information(self, "ΟΚ", "Η ρίζα φακέλων αποθηκεύτηκε.")

    def on_has_taxis_changed(self, txt: str):
        self.taxis_box.setVisible(txt == "Ναι")

    def refresh_next_id(self):
        self.lbl_next_id.setText(f"Επόμενο ID: {get_next_id()}")
        self.lbl_date.setText(date.today().isoformat())

    def selected_services(self) -> List[Tuple[str, int]]:
        selected = []
        for cb, (name, price) in zip(self.service_checks, SERVICES):
            if cb.isChecked():
                selected.append((name, price))
        return selected

    def recalculate(self):
        total = sum(price for _, price in self.selected_services())
        paid = safe_float(self.in_paid.text())
        balance = total - paid
        self.lbl_total.setText(f"Σύνολο: {total:.2f}")
        self.lbl_balance.setText(f"Υπόλοιπο: {balance:.2f}")

    def save_customer(self):
        name = self.in_name.text().strip()
        phone = self.in_phone.text().strip()
        afm = self.in_afm.text().strip()
        amka = self.in_amka.text().strip()
        ama = self.in_ama.text().strip()

        if not name:
            QMessageBox.warning(self, "Λείπει κάτι", "Το ονοματεπώνυμο είναι υποχρεωτικό.")
            return

        if afm and len(afm) != 9:
            QMessageBox.warning(self, "Σφάλμα", "Το ΑΦΜ πρέπει να έχει 9 ψηφία (ή να μείνει κενό).")
            return

        if amka and len(amka) != 11:
            QMessageBox.warning(self, "Σφάλμα", "Το ΑΜΚΑ πρέπει να έχει 11 ψηφία (ή να μείνει κενό).")
            return

        services = self.selected_services()
        if not services:
            QMessageBox.warning(self, "Λείπει κάτι", "Διάλεξε τουλάχιστον 1 ενέργεια/υπηρεσία.")
            return

        if not self.cb_confirm_data.isChecked() or not self.cb_confirm_services.isChecked():
            QMessageBox.warning(self, "Απαιτείται επιβεβαίωση", "Πρέπει να επιβεβαιώσεις στοιχεία και υπηρεσίες πριν την αποθήκευση.")
            return

        total = float(sum(p for _, p in services))
        paid = safe_float(self.in_paid.text())
        balance = total - paid

        customer_id = get_next_id()
        folder = ensure_customer_folder(customer_id, name)

        has_taxis = "Yes" if self.cb_has_taxis.currentText() == "Ναι" else "No"
        t_user = self.in_taxis_user.text().strip() if has_taxis == "Yes" else ""
        t_pass = self.in_taxis_pass.text() if has_taxis == "Yes" else ""

        kleid = "Confirmed" if self.in_kleid.currentText() == "Επιβεβαιώθηκε" else "Rejected"
        amka_status = "Confirmed" if self.in_amka_status.currentText() == "Επιβεβαιώθηκε" else "Rejected"
        aporr = "Yes" if self.in_aporr.currentText() == "Ναι" else "No"

        wb, ws = open_ws()

        record = {
            "ID": customer_id,
            "ServiceType": "KEP" if self.service_type.currentText() == "ΚΕΠ" else "Nomika",
            "Date": date.today().isoformat(),
            "Name": name,
            "Phone": phone,
            "AFM": afm,
            "AMKA": amka,
            "AMA": ama,
            "HasTaxisnet": has_taxis,
            "TaxisnetUser": t_user,
            "TaxisnetPass": t_pass,
            "Kleidarithmos": kleid,
            "AMKA_AMA": amka_status,
            "Aporipsi": aporr,
            "ActionsToday": ", ".join([f"{n}({p})" for n, p in services]),
            "PaymentMethod": self.cmb_payment.currentText(),
            "RequestNotes": self.in_request.toPlainText(),
            "Total": float(total),
            "Paid": float(paid),
            "Balance": float(balance),
            "FolderPath": str(folder),
            "FilesConfirmedBy": "",
            "FilesConfirmedAt": "",
            "CreatedBy": self.user,
            "CreatedAt": now_iso(),
            "LastEditedBy": self.user,
            "LastEditedAt": now_iso(),
            # Νέα πεδία με default τιμές
            "Goal": "",
            "DeclarationStatus": "Μη Ορισμένη",
            "CustomerStatus": "Νέος",
            "Amount": str(float(total)) if total else "",
            "AMKA_Valid": "Άγνωστο",
        }

        append_record_by_headers(ws, record)
        wb.save(EXCEL_PATH)

        audit_log(self.user, "CREATE", customer_id, f"Σύνολο={total}, Πληρώθηκε={paid}, Υπόλοιπο={balance}")

        self.last_saved_customer_id = customer_id
        self.last_saved_customer_folder = folder

        self.refresh_next_id()
        self.recalculate()
        self.refresh_file_lists()
        self.refresh_dashboard()

        QMessageBox.information(self, "ΟΚ", f"Ο πελάτης αποθηκεύτηκε με ID {customer_id}.")
        self.clear_form()

    def clear_form(self):
        self.in_name.clear()
        self.in_phone.clear()
        self.in_afm.clear()
        self.in_amka.clear()
        self.in_ama.clear()
        self.cb_has_taxis.setCurrentIndex(0)
        self.in_taxis_user.clear()
        self.in_taxis_pass.clear()
        self.in_request.clear()
        self.in_paid.clear()
        self.cmb_payment.setCurrentIndex(0)
        for cb in self.service_checks:
            cb.setChecked(False)
        self.cb_confirm_data.setChecked(False)
        self.cb_confirm_services.setChecked(False)
        self.pending_files = []
        self.list_pending.clear()
        self.recalculate()
        self.refresh_next_id()

    def open_last_folder(self):
        if not self.last_saved_customer_folder:
            QMessageBox.information(self, "Πληροφορία", "Δεν υπάρχει αποθηκευμένος πελάτης ακόμη.")
            return
        if not self.last_saved_customer_folder.exists():
            QMessageBox.warning(self, "Σφάλμα", "Ο φάκελος δεν υπάρχει.")
            return
        open_in_explorer(self.last_saved_customer_folder)

    # ---------------- Files tab actions ----------------
    def on_files_dropped(self, paths: List[str]):
        self.pending_files.extend(paths)
        self.refresh_pending_list()

    def pick_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Επιλογή αρχείων")
        if files:
            self.pending_files.extend(files)
            self.refresh_pending_list()

    def refresh_pending_list(self):
        self.list_pending.clear()
        for p in self.pending_files:
            self.list_pending.addItem(QListWidgetItem(p))

    def set_scanner_folder_ui(self):
        set_scanner_folder(self.in_scanner.text().strip())
        QMessageBox.information(self, "ΟΚ", "Ο φάκελος scanner αποθηκεύτηκε.")
        self.refresh_file_lists()

    def browse_scanner_folder(self):
        start_dir = self.in_scanner.text().strip()
        if not start_dir or not Path(start_dir).exists():
            start_dir = str(Path.home())

        folder = QFileDialog.getExistingDirectory(self, "Επίλεξε φάκελο scanner", start_dir)
        if folder:
            self.in_scanner.setText(folder)
            set_scanner_folder(folder)
            self.refresh_file_lists()

    def browse_scanner_file(self):
        start_dir = self.in_scanner.text().strip()
        if not start_dir or not Path(start_dir).exists():
            start_dir = str(Path.home())

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Επίλεξε ένα αρχείο από τον scanner",
            start_dir,
            "Όλα τα αρχεία (*.*)"
        )
        if file_path:
            folder = str(Path(file_path).parent)
            self.in_scanner.setText(folder)
            set_scanner_folder(folder)
            self.refresh_file_lists()

            filename = Path(file_path).name
            for i in range(self.list_scanner.count()):
                if self.list_scanner.item(i).text() == filename:
                    self.list_scanner.setCurrentRow(i)
                    break

    def refresh_file_lists(self):
        self.list_folder.clear()
        if self.last_saved_customer_folder and self.last_saved_customer_folder.exists():
            for p in sorted(self.last_saved_customer_folder.iterdir()):
                if p.is_file():
                    self.list_folder.addItem(QListWidgetItem(p.name))

        self.list_scanner.clear()
        sp = self.in_scanner.text().strip()
        if sp:
            sf = Path(sp)
            if sf.exists():
                for p in sorted(sf.iterdir()):
                    if p.is_file():
                        self.list_scanner.addItem(QListWidgetItem(p.name))

    def move_pending_to_customer(self):
        if not self.last_saved_customer_folder or self.last_saved_customer_id is None:
            QMessageBox.warning(self, "Λείπει κάτι", "Πρώτα αποθήκευσε πελάτη.")
            return
        if not self.pending_files:
            QMessageBox.information(self, "Πληροφορία", "Δεν υπάρχουν pending αρχεία.")
            return

        moved = 0
        for p in list(self.pending_files):
            src = Path(p)
            if not src.exists():
                continue
            dst = self.last_saved_customer_folder / src.name
            if dst.exists():
                stem, suf = dst.stem, dst.suffix
                i = 1
                while (self.last_saved_customer_folder / f"{stem}_{i}{suf}").exists():
                    i += 1
                dst = self.last_saved_customer_folder / f"{stem}_{i}{suf}"
            try:
                src.replace(dst)
                moved += 1
            except Exception:
                pass

        self.pending_files = []
        self.refresh_pending_list()
        self.refresh_file_lists()
        audit_log(self.user, "ADD_FILES", self.last_saved_customer_id, f"moved_pending={moved}")
        QMessageBox.information(self, "ΟΚ", f"Μεταφέρθηκαν {moved} αρχεία.")

    def move_selected_scanner_to_customer(self):
        if not self.last_saved_customer_folder or self.last_saved_customer_id is None:
            QMessageBox.warning(self, "Λείπει κάτι", "Πρώτα αποθήκευσε πελάτη.")
            return
        sp = self.in_scanner.text().strip()
        if not sp:
            QMessageBox.warning(self, "Λείπει κάτι", "Θέσε φάκελο scanner πρώτα.")
            return
        sf = Path(sp)
        if not sf.exists():
            QMessageBox.warning(self, "Σφάλμα", "Ο φάκελος scanner δεν υπάρχει.")
            return

        item = self.list_scanner.currentItem()
        if not item:
            QMessageBox.information(self, "Πληροφορία", "Επίλεξε ένα αρχείο από τον scanner.")
            return

        src = sf / item.text()
        if not src.exists():
            return

        dst = self.last_saved_customer_folder / src.name
        if dst.exists():
            stem, suf = dst.stem, dst.suffix
            i = 1
            while (self.last_saved_customer_folder / f"{stem}_{i}{suf}").exists():
                i += 1
            dst = self.last_saved_customer_folder / f"{stem}_{i}{suf}"

        try:
            src.replace(dst)
            audit_log(self.user, "ADD_FILES", self.last_saved_customer_id, f"moved_scanner=1 file={dst.name}")
            self.refresh_file_lists()
            QMessageBox.information(self, "ΟΚ", "Το αρχείο μεταφέρθηκε.")
        except Exception as e:
            QMessageBox.warning(self, "Σφάλμα", str(e))

    def print_selected_customer_file(self):
        if not self.last_saved_customer_folder or not self.last_saved_customer_folder.exists():
            return
        item = self.list_folder.currentItem()
        if not item:
            return
        p = self.last_saved_customer_folder / item.text()
        if not p.exists():
            return
        ok = print_with_dialog(p, self)
        if not ok:
            QMessageBox.information(self, "Εκτύπωση", "Αποτυχία εκτύπωσης. Θα ανοίξω το αρχείο.")
            open_in_explorer(p)
        if self.last_saved_customer_id is not None:
            audit_log(self.user, "PRINT", self.last_saved_customer_id, p.name)

    def confirm_files(self):
        if self.last_saved_customer_id is None:
            QMessageBox.warning(self, "Λείπει κάτι", "Πρώτα αποθήκευσε πελάτη.")
            return

        wb, ws = open_ws()
        cols = headers_map(ws)
        for r in range(2, ws.max_row + 1):
            if str(ws.cell(r, cols.get("ID", 1)).value).strip() == str(self.last_saved_customer_id):
                update_row(ws, r, {
                    "FilesConfirmedBy": self.user,
                    "FilesConfirmedAt": now_iso(),
                    "LastEditedBy": self.user,
                    "LastEditedAt": now_iso(),
                })
                wb.save(EXCEL_PATH)
                audit_log(self.user, "CONFIRM_FILES", self.last_saved_customer_id, "Επιβεβαιώθηκαν αρχεία")
                QMessageBox.information(self, "ΟΚ", "Τα αρχεία επιβεβαιώθηκαν.")
                self.refresh_dashboard()
                return

        QMessageBox.warning(self, "Σφάλμα", "Ο πελάτης δεν βρέθηκε στο Excel.")

    # ---------------- Tab: Αναζήτηση ----------------
    def build_tab_search(self, parent: QWidget):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setSpacing(10)

        search_card = QFrame()
        search_card.setObjectName("card")
        c = QVBoxLayout(search_card)
        c.setContentsMargins(12, 12, 12, 12)
        c.setSpacing(8)

        row = QHBoxLayout()
        row.setSpacing(8)
        self.q_id = QLineEdit(); self.q_id.setPlaceholderText("ID")
        self.q_name = QLineEdit(); self.q_name.setPlaceholderText("Όνομα περιέχει…")
        self.q_phone = QLineEdit(); self.q_phone.setPlaceholderText("Τηλέφωνο περιέχει…")
        btn = QPushButton("Αναζήτηση")
        btn.clicked.connect(self.run_search)

        row.addWidget(self.q_id, 1)
        row.addWidget(self.q_name, 2)
        row.addWidget(self.q_phone, 2)
        row.addWidget(btn, 1)
        c.addLayout(row)

        split = QSplitter(Qt.Horizontal)
        split.setChildrenCollapsible(False)
        split.setHandleWidth(8)

        left = QFrame()
        left.setObjectName("subcard")
        left_l = QVBoxLayout(left)
        left_l.setContentsMargins(10, 10, 10, 10)
        left_l.setSpacing(6)

        self.table = QTableWidget(0, 7)
        self.table.setHorizontalHeaderLabels(["Γραμμή", "ID", "Όνομα", "Τηλέφωνο", "Υπόλοιπο", "Αρχεία;", "Ηλικία"])
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.SingleSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeToContents)
        polish_table(self.table)

        self.table.cellDoubleClicked.connect(self.open_selected_customer_dialog)
        self.table.itemSelectionChanged.connect(self.preview_selected_customer)
        left_l.addWidget(self.table)

        left_btns = QHBoxLayout()
        left_btns.setSpacing(6)
        open_btn = QPushButton("Άνοιγμα καρτέλας")
        open_btn.setObjectName("secondary")
        open_btn.clicked.connect(self.open_selected_customer_dialog)

        refresh_btn = QPushButton("Ανανέωση προεπισκόπησης")
        refresh_btn.setObjectName("secondary")
        refresh_btn.clicked.connect(self.preview_selected_customer)

        left_btns.addWidget(open_btn)
        left_btns.addWidget(refresh_btn)
        left_btns.addStretch(1)
        left_l.addLayout(left_btns)

        self.preview_card = CustomerCardWidget()
        self.preview_card.bind_actions(
            on_open_folder=self.search_open_folder,
            on_open_selected=self.search_open_selected_file,
            on_print_selected=self.search_print_selected_file
        )

        split.addWidget(left)
        split.addWidget(self.preview_card)
        split.setSizes([600, 400])

        c.addWidget(split)
        lay.addWidget(search_card)
        self.preview_card.clear()

    def run_search(self):
        qid = self.q_id.text().strip()
        qn = self.q_name.text().strip()
        qp = self.q_phone.text().strip()

        if not (qid or qn or qp):
            QMessageBox.warning(self, "Λείπει κάτι", "Συμπλήρωσε τουλάχιστον 1 πεδίο αναζήτησης.")
            return

        rows = find_rows(query_id=qid, query_name=qn, query_phone=qp)
        wb, ws = open_ws()
        self.table.setRowCount(0)

        for r in rows:
            rec = row_to_record(ws, r)
            cid = int(rec.get("ID") or 0)
            name = rec.get("Name")
            phone = rec.get("Phone")
            bal = rec.get("Balance")
            folder_raw = str(rec.get("FolderPath") or "").strip()
            folder = Path(folder_raw) if folder_raw else None
            has_files = "Ναι" if (folder and folder.exists() and any(p.is_file() for p in folder.iterdir())) else "Όχι"
            age = human_duration(str(rec.get("CreatedAt") or ""))

            row_i = self.table.rowCount()
            self.table.insertRow(row_i)
            self.table.setItem(row_i, 0, QTableWidgetItem(str(r)))
            self.table.setItem(row_i, 1, QTableWidgetItem(str(cid)))
            self.table.setItem(row_i, 2, QTableWidgetItem(str(name)))
            self.table.setItem(row_i, 3, QTableWidgetItem(str(phone)))
            self.table.setItem(row_i, 4, QTableWidgetItem(str(bal)))
            self.table.setItem(row_i, 5, QTableWidgetItem(has_files))
            self.table.setItem(row_i, 6, QTableWidgetItem(age))

        audit_log(self.user, "SEARCH", details=f"id={qid}, name={qn}, phone={qp}")

        if self.table.rowCount() > 0:
            self.table.selectRow(0)
            self.preview_selected_customer()
        else:
            self.preview_card.clear()

    def selected_excel_row_from_table(self) -> Optional[int]:
        row = self.table.currentRow()
        if row < 0:
            return None
        try:
            return int(self.table.item(row, 0).text())
        except Exception:
            return None

    def preview_selected_customer(self):
        excel_row = self.selected_excel_row_from_table()
        if not excel_row:
            self.preview_card.clear()
            return
        wb, ws = open_ws()
        rec = row_to_record(ws, excel_row)
        self.search_selected_excel_row = excel_row
        self.preview_card.set_record(excel_row, rec)

    def open_selected_customer_dialog(self, *_):
        excel_row = self.selected_excel_row_from_table()
        if not excel_row:
            return
        dlg = CustomerDialog(self.user, excel_row)
        dlg.exec()
        self.preview_selected_customer()
        self.refresh_dashboard()

    def search_open_folder(self):
        if not self.preview_card.folder_path:
            QMessageBox.information(self, "Πληροφορία", "Δεν υπάρχει αποθηκευμένος φάκελος.")
            return
        if not self.preview_card.folder_path.exists():
            QMessageBox.warning(self, "Σφάλμα", "Ο φάκελος δεν υπάρχει στο δίσκο.")
            return
        open_in_explorer(self.preview_card.folder_path)

    def search_open_selected_file(self):
        p = self.preview_card.selected_file_path()
        if not p:
            return
        open_in_explorer(p)

    def search_print_selected_file(self):
        p = self.preview_card.selected_file_path()
        if not p:
            return
        ok = print_with_dialog(p, self)
        if not ok:
            QMessageBox.information(self, "Εκτύπωση", "Δεν μπόρεσα να κάνω άμεση εκτύπωση. Θα ανοίξω το αρχείο.")
            open_in_explorer(p)
        if self.preview_card.customer_id is not None:
            audit_log(self.user, "PRINT", self.preview_card.customer_id, p.name)

    # ---------------- Tab: Dashboard ----------------
    def build_tab_dashboard(self, parent: QWidget):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setSpacing(10)

        card = QFrame()
        card.setObjectName("card")
        c = QVBoxLayout(card)
        c.setContentsMargins(12, 12, 12, 12)
        c.setSpacing(8)

        filters_grid = QGridLayout()
        filters_grid.setHorizontalSpacing(10)
        filters_grid.setVerticalSpacing(8)

        self.f_id = QLineEdit(); self.f_id.setPlaceholderText("ID (π.χ. 4 ή 4-10)")
        self.f_name = QLineEdit(); self.f_name.setPlaceholderText("Όνομα περιέχει")
        self.f_phone = QLineEdit(); self.f_phone.setPlaceholderText("Τηλέφωνο περιέχει")

        self.f_priority = QComboBox(); self.f_priority.addItems(["Όλα", "Έχει υπόλοιπο", "Χωρίς υπόλοιπο"])
        self.f_serving = QComboBox(); self.f_serving.addItems(["Όλα", "Σήμερα", "Τελευταίο 24ωρο", "Τελευταίες 7 ημέρες", "Πάνω από 7 ημέρες"])
        self.f_sort_id = QComboBox(); self.f_sort_id.addItems(["ID ↑ (μικρό→μεγάλο)", "ID ↓ (μεγάλο→μικρό)"])

        filters_grid.addWidget(QLabel("ID:"), 0, 0)
        filters_grid.addWidget(self.f_id, 0, 1)
        filters_grid.addWidget(QLabel("Όνομα:"), 0, 2)
        filters_grid.addWidget(self.f_name, 0, 3)
        filters_grid.addWidget(QLabel("Τηλέφωνο:"), 0, 4)
        filters_grid.addWidget(self.f_phone, 0, 5)

        filters_grid.addWidget(QLabel("Υπόλοιπο:"), 1, 0)
        filters_grid.addWidget(self.f_priority, 1, 1)
        filters_grid.addWidget(QLabel("Χρονικό διάστημα:"), 1, 2)
        filters_grid.addWidget(self.f_serving, 1, 3)
        filters_grid.addWidget(QLabel("Ταξινόμηση:"), 1, 4)
        filters_grid.addWidget(self.f_sort_id, 1, 5)

        c.addLayout(filters_grid)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        btn_apply = QPushButton("Εφαρμογή φίλτρων")
        btn_apply.clicked.connect(self.refresh_dashboard)

        btn_clear = QPushButton("Καθαρισμός")
        btn_clear.setObjectName("secondary")
        btn_clear.clicked.connect(self.clear_dashboard_filters)

        btn_pdf = QPushButton("Εξαγωγή → PDF")
        btn_pdf.setObjectName("secondary")
        btn_pdf.clicked.connect(self.export_dashboard_to_pdf)

        btn_row.addWidget(btn_apply)
        btn_row.addWidget(btn_clear)
        btn_row.addWidget(btn_pdf)
        btn_row.addStretch(1)
        c.addLayout(btn_row)

        lay.addWidget(card)

        self.dash = QTableWidget(0, 6)
        self.dash.setHorizontalHeaderLabels(["ID", "Όνομα", "Τηλέφωνο", "Υπόλοιπο", "Αρχεία", "Χρόνος"])
        self.dash.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.dash.setEditTriggers(QAbstractItemView.NoEditTriggers)

        self.dash.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.dash.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.dash.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        self.dash.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        self.dash.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeToContents)
        self.dash.horizontalHeader().setSectionResizeMode(5, QHeaderView.ResizeToContents)
        polish_table(self.dash)

        lay.addWidget(self.dash, 1)

    def clear_dashboard_filters(self):
        self.f_id.clear()
        self.f_name.clear()
        self.f_phone.clear()
        self.f_priority.setCurrentIndex(0)
        self.f_serving.setCurrentIndex(0)
        self.f_sort_id.setCurrentIndex(0)
        self.refresh_dashboard()

    def export_dashboard_to_pdf(self):
        path, _ = QFileDialog.getSaveFileName(self, "Εξαγωγή πίνακα σε PDF", "dashboard.pdf", "PDF (*.pdf)")
        if not path:
            return

        try:
            printer = QPrinter(QPrinter.HighResolution)
            printer.setOutputFormat(QPrinter.PdfFormat)
            printer.setOutputFileName(path)

            page_layout = QPageLayout(
                QPageSize(QPageSize.A4),
                QPageLayout.Landscape,
                QMarginsF(15, 15, 15, 15)
            )
            printer.setPageLayout(page_layout)

            doc_html = []
            doc_html.append("<html><head><meta charset='utf-8'><style>")
            doc_html.append("body{font-family:Arial; font-size:10pt;}")
            doc_html.append("table{border-collapse:collapse; width:100%;}")
            doc_html.append("th{background:#f0f0f0; padding:6px; border:1px solid #999; text-align:left;}")
            doc_html.append("td{padding:6px; border:1px solid #ccc;}")
            doc_html.append("tr:nth-child(even){background:#fafafa;}")
            doc_html.append("</style></head><body>")
            doc_html.append("<h2>Client Manager – Εξαγωγή Πίνακα</h2>")
            doc_html.append(f"<div>Ημερομηνία εξαγωγής: {_html.escape(now_iso())}</div><br>")
            doc_html.append("<table><tr>")

            headers = ["ID", "Όνομα", "Τηλέφωνο", "Υπόλοιπο", "Αρχεία", "Χρόνος"]
            for h in headers:
                doc_html.append(f"<th>{_html.escape(h)}</th>")
            doc_html.append("</tr>")

            for row in range(self.dash.rowCount()):
                doc_html.append("<tr>")
                for col in range(self.dash.columnCount()):
                    item = self.dash.item(row, col)
                    val = item.text() if item else ""
                    doc_html.append(f"<td>{_html.escape(val)}</td>")
                doc_html.append("</tr>")

            doc_html.append("</table></body></html>")

            doc = QTextDocument()
            doc.setHtml("".join(doc_html))

            page_rect = printer.pageRect(QPrinter.Point)  # ✅ fixed
            doc.setPageSize(page_rect.size())
            doc.print_(printer)

            QMessageBox.information(self, "ΟΚ", f"Το PDF αποθηκεύτηκε:\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Σφάλμα PDF", f"{e}")

    def refresh_dashboard(self):
        wb, ws = open_ws()

        fid = self.f_id.text().strip()
        fname = self.f_name.text().strip().lower()
        fphone = self.f_phone.text().strip()
        fpriority = self.f_priority.currentText()
        fserv = self.f_serving.currentText()
        sort_mode = self.f_sort_id.currentText()

        items = []

        def id_allowed(cid_int: int) -> bool:
            if not fid:
                return True
            s = fid.replace(" ", "")
            if "-" in s:
                try:
                    a, b = s.split("-", 1)
                    a = int(a); b = int(b)
                    lo, hi = min(a, b), max(a, b)
                    return lo <= cid_int <= hi
                except Exception:
                    return False
            try:
                return cid_int == int(s)
            except Exception:
                return False

        for r in range(2, ws.max_row + 1):
            rec = row_to_record(ws, r)

            try:
                cid_int = int(rec.get("ID") or 0)
            except Exception:
                cid_int = 0

            name = str(rec.get("Name") or "")
            phone = str(rec.get("Phone") or "")
            bal = safe_float(str(rec.get("Balance") or "0"))
            created = str(rec.get("CreatedAt") or "")
            folder_raw = str(rec.get("FolderPath") or "").strip()
            folder = Path(folder_raw) if folder_raw else None

            if not id_allowed(cid_int):
                continue
            if fname and fname not in name.lower():
                continue
            if fphone and fphone not in phone:
                continue

            if fpriority == "Έχει υπόλοιπο" and bal <= 0:
                continue
            if fpriority == "Χωρίς υπόλοιπο" and bal > 0:
                continue

            if fserv != "Όλα":
                try:
                    dt = datetime.fromisoformat(created)
                    diff = datetime.now() - dt
                    if fserv == "Σήμερα" and diff.days != 0:
                        continue
                    if fserv == "Τελευταίο 24ωρο" and diff.total_seconds() > 86400:
                        continue
                    if fserv == "Τελευταίες 7 ημέρες" and diff.days > 7:
                        continue
                    if fserv == "Πάνω από 7 ημέρες" and diff.days <= 7:
                        continue
                except Exception:
                    continue

            files = "Ναι" if (folder and folder.exists() and any(p.is_file() for p in folder.iterdir())) else "Όχι"
            serving_time = human_duration(created)
            items.append((cid_int, name, phone, bal, files, serving_time))

        reverse = (sort_mode == "ID ↓ (μεγάλο→μικρό)")
        items.sort(key=lambda x: x[0], reverse=reverse)

        self.dash.setRowCount(0)
        for cid_int, name, phone, bal, files, serving_time in items:
            row_i = self.dash.rowCount()
            self.dash.insertRow(row_i)
            self.dash.setItem(row_i, 0, QTableWidgetItem(str(cid_int)))
            self.dash.setItem(row_i, 1, QTableWidgetItem(name))
            self.dash.setItem(row_i, 2, QTableWidgetItem(phone))
            self.dash.setItem(row_i, 3, QTableWidgetItem(f"{bal:.2f}"))
            self.dash.setItem(row_i, 4, QTableWidgetItem(files))
            self.dash.setItem(row_i, 5, QTableWidgetItem(serving_time))
    # ---------------- Tab: Settings ----------------
    def build_tab_settings(self, parent: QWidget):
        lay = QVBoxLayout(parent)
        lay.setContentsMargins(8, 8, 8, 8)
        lay.setSpacing(10)

        card = QFrame()
        card.setObjectName("card")
        c = QVBoxLayout(card)
        c.setContentsMargins(12, 12, 12, 12)
        c.setSpacing(10)

        title = QLabel("Ρυθμίσεις Εφαρμογής")
        title.setObjectName("title")
        c.addWidget(title)

        info = QLabel(
            "Εδώ αλλάζεις Paths & defaults.\n"
            "⚠️ Αν αλλάξεις Data Root (Excel/users/audit), χρειάζεται επανεκκίνηση για να φορτώσει τα νέα αρχεία."
        )
        info.setObjectName("muted")
        info.setWordWrap(True)
        c.addWidget(info)

        # -------- Inputs --------
        self.set_data_root_in = QLineEdit(get_data_root())
        self.set_data_root_in.setPlaceholderText(
            r"Data Root (π.χ. D:\ClientManagerV1 ή \\SERVER\Share\ClientManagerV1)"
        )

        self.set_clients_root_in = QLineEdit(str(get_clients_root()))
        self.set_clients_root_in.setPlaceholderText(r"Clients Root (π.χ. D:\Clients)")

        self.set_scanner_in = QLineEdit(get_scanner_folder())
        self.set_scanner_in.setPlaceholderText(r"Φάκελος Scanner (π.χ. C:\Scanner)")

        self.set_theme_combo = QComboBox()
        self.set_theme_combo.addItems(["light", "dark"])
        self.set_theme_combo.setCurrentText(get_theme())

        grid = QGridLayout()
        grid.setHorizontalSpacing(10)
        grid.setVerticalSpacing(8)

        # Data Root row
        grid.addWidget(QLabel("Data Root (Excel/users/audit):"), 0, 0)
        grid.addWidget(self.set_data_root_in, 0, 1)

        btn_pick_data = QPushButton("Επιλογή…")
        btn_pick_data.setObjectName("secondary")
        btn_pick_data.clicked.connect(self.pick_data_root)
        grid.addWidget(btn_pick_data, 0, 2)

        # Clients Root row
        grid.addWidget(QLabel("Clients Root (φάκελοι πελατών):"), 1, 0)
        grid.addWidget(self.set_clients_root_in, 1, 1)

        btn_pick_clients = QPushButton("Επιλογή…")
        btn_pick_clients.setObjectName("secondary")
        btn_pick_clients.clicked.connect(self.pick_clients_root_from_settings)
        grid.addWidget(btn_pick_clients, 1, 2)

        # Scanner row
        grid.addWidget(QLabel("Scanner folder:"), 2, 0)
        grid.addWidget(self.set_scanner_in, 2, 1)

        btn_pick_scanner = QPushButton("Επιλογή…")
        btn_pick_scanner.setObjectName("secondary")
        btn_pick_scanner.clicked.connect(self.pick_scanner_folder_from_settings)
        grid.addWidget(btn_pick_scanner, 2, 2)

        # Theme row
        grid.addWidget(QLabel("Theme:"), 3, 0)
        grid.addWidget(self.set_theme_combo, 3, 1)

        c.addLayout(grid)

        # Buttons
        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)

        btn_apply = QPushButton("Αποθήκευση / Εφαρμογή")
        btn_apply.clicked.connect(self.apply_settings_from_tab)

        btn_refresh = QPushButton("Ανανέωση εμφάνισης paths")
        btn_refresh.setObjectName("secondary")
        btn_refresh.clicked.connect(self.show_current_paths)

        btn_row.addWidget(btn_apply)
        btn_row.addWidget(btn_refresh)
        btn_row.addStretch(1)
        c.addLayout(btn_row)

        # Current paths box
        c.addWidget(QLabel("Τρέχοντα paths (ό,τι χρησιμοποιεί τώρα η εφαρμογή):"))
        self.lbl_roots = QTextEdit()
        self.lbl_roots.setReadOnly(True)
        self.lbl_roots.setMinimumHeight(160)
        c.addWidget(self.lbl_roots)

        lay.addWidget(card)
        lay.addStretch(1)

        self.show_current_paths()

    def show_current_paths(self):
        txt = []
        txt.append(f"DATA_DIR: {DATA_DIR}")
        txt.append(f"EXCEL_PATH: {EXCEL_PATH}")
        txt.append(f"USERS_PATH: {USERS_PATH}")
        txt.append(f"AUDIT_PATH: {AUDIT_PATH}")
        txt.append("")
        txt.append(f"Saved setting data_root: {get_data_root() or '(default roaming)'}")
        txt.append(f"Saved setting clients_root: {get_clients_root()}")
        txt.append(f"Saved setting scanner_folder: {get_scanner_folder() or '(empty)'}")
        txt.append(f"Saved setting theme: {get_theme()}")
        self.lbl_roots.setPlainText("\n".join(txt))

    # -------- Settings helpers (MISSING before) --------
    def pick_data_root(self):
        start_dir = self.set_data_root_in.text().strip()
        if not start_dir or not Path(start_dir).exists():
            start_dir = str(Path.home())

        folder = QFileDialog.getExistingDirectory(self, "Επίλεξε Data Root", start_dir)
        if folder:
            self.set_data_root_in.setText(folder)

    def pick_clients_root_from_settings(self):
        start_dir = self.set_clients_root_in.text().strip()
        if not start_dir or not Path(start_dir).exists():
            start_dir = str(Path.home())

        folder = QFileDialog.getExistingDirectory(self, "Επίλεξε Clients Root", start_dir)
        if folder:
            self.set_clients_root_in.setText(folder)

    def pick_scanner_folder_from_settings(self):
        start_dir = self.set_scanner_in.text().strip()
        if not start_dir or not Path(start_dir).exists():
            start_dir = str(Path.home())

        folder = QFileDialog.getExistingDirectory(self, "Επίλεξε φάκελο Scanner", start_dir)
        if folder:
            self.set_scanner_in.setText(folder)

    def apply_settings_from_tab(self):
        # 1) Data Root (requires restart because module globals were built at startup)
        new_data_root = self.set_data_root_in.text().strip()
        old_data_root = (get_data_root() or "").strip()

        # 2) Clients Root (can apply immediately)
        new_clients_root = self.set_clients_root_in.text().strip()

        # 3) Scanner folder (can apply immediately)
        new_scanner = self.set_scanner_in.text().strip()

        # 4) Theme (can apply immediately)
        new_theme = self.set_theme_combo.currentText().strip()

        # Validate/create folders where applicable
        try:
            if new_data_root:
                Path(new_data_root).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            QMessageBox.warning(self, "Σφάλμα", f"Δεν μπορώ να χρησιμοποιήσω το Data Root:\n{e}")
            return

        try:
            if new_clients_root:
                Path(new_clients_root).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            QMessageBox.warning(self, "Σφάλμα", f"Δεν μπορώ να χρησιμοποιήσω το Clients Root:\n{e}")
            return

        try:
            if new_scanner:
                Path(new_scanner).mkdir(parents=True, exist_ok=True)
        except Exception as e:
            QMessageBox.warning(self, "Σφάλμα", f"Δεν μπορώ να χρησιμοποιήσω το Scanner folder:\n{e}")
            return

        # Save settings
        if new_data_root != old_data_root:
            set_data_root(new_data_root)

        if new_clients_root:
            set_clients_root(new_clients_root)
            # sync also the New tab input if it exists
            if hasattr(self, "in_clients_root"):
                self.in_clients_root.setText(new_clients_root)

        set_scanner_folder(new_scanner)
        if hasattr(self, "in_scanner"):
            self.in_scanner.setText(new_scanner)

        # Apply theme now
        if new_theme in ("light", "dark"):
            self.apply_theme(new_theme)

        # Update UI text
        self.show_current_paths()
        self.refresh_file_lists()

        # Restart warning only if data_root changed
        if new_data_root != old_data_root:
            QMessageBox.information(
                self,
                "ΟΚ",
                "Οι ρυθμίσεις αποθηκεύτηκαν.\n\n"
                "⚠️ Έχει αλλάξει το Data Root, οπότε χρειάζεται επανεκκίνηση της εφαρμογής\n"
                "για να ενημερωθούν τα paths (Excel/users/audit)."
            )
        else:
            QMessageBox.information(self, "ΟΚ", "Οι ρυθμίσεις αποθηκεύτηκαν και εφαρμόστηκαν.")

    
        
# ----------------------------- Global Error Hook (για EXE) -----------------------------
def excepthook(exc_type, exc, tb):
    msg = "".join(traceback.format_exception(exc_type, exc, tb))
    try:
        QMessageBox.critical(None, "Κρίσιμο σφάλμα", msg)
    except Exception:
        pass
    sys.__stderr__.write(msg + "\n")

sys.excepthook = excepthook


# ----------------------------- Main -----------------------------
def main():
    ensure_excel()
    ensure_users_file()

    app = QApplication(sys.argv)
    app.setStyleSheet(LIGHT_STYLE if get_theme() == "light" else DARK_STYLE)

    w = LoginWindow()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
